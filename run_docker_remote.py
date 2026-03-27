# -*- coding: utf-8 -*-
"""Ejecutar proyecto POD en Docker remoto y validar logs/errores."""
import paramiko
import sys
import time
import os

# Evitar UnicodeEncodeError en consola Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

HOST = "20.102.51.145"
USER = "pdiaz"
PASS = "dF62(5PR7U54"

def run_ssh(client, cmd, timeout=60):
    stdin, stdout, stderr = client.exec_command(cmd, timeout=timeout)
    out = stdout.read().decode("utf-8", errors="replace")
    err = stderr.read().decode("utf-8", errors="replace")
    return out, err

def main():
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        print(f"Conectando a {USER}@{HOST}...")
        client.connect(HOST, username=USER, password=PASS, timeout=15)
        print("[OK] Conectado.\n")

        # 1. Ver estado de Docker
        print("=== 1. IMÁGENES Y CONTENEDORES ===")
        out, _ = run_ssh(client, "docker images | grep -E 'padua|POD|REPOSITORY'")
        print(out or "(ninguna imagen padua)")
        out, _ = run_ssh(client, "docker ps -a")
        print(out)

        # 2. Verificar si existe imagen padua-pod
        out, _ = run_ssh(client, "docker images -q padua-pod:latest")
        if not out.strip():
            print("\n[WARNING] Imagen padua-pod:latest no encontrada.")
            print("Listando todas las imágenes:")
            out, _ = run_ssh(client, "docker images")
            print(out)
            client.close()
            return

        # 3. Detener y eliminar contenedor anterior si existe
        print("\n=== 2. LIMPIAR CONTENEDOR ANTERIOR ===")
        run_ssh(client, "docker stop padua-pod-run 2>/dev/null; docker rm padua-pod-run 2>/dev/null; echo 'OK'")

        # 4. Crear carpetas y GUIAS.xlsx con las guías indicadas
        GUIAS = ["111105823079", "111105822097", "888005818713", "888005786000",
                 "111105820570", "888005805109", "111105820512", "888005746029"]
        run_ssh(client, "mkdir -p /home/pdiaz/WEBSITE /home/pdiaz/DESCARGA")
        print(f"   Actualizando GUIAS.xlsx con {len(GUIAS)} guías: {', '.join(GUIAS)}")
        guias_repr = repr(GUIAS)
        out2, _ = run_ssh(client, f'''docker run --rm --entrypoint python3 -v /home/pdiaz/WEBSITE:/data padua-pod:latest -c "
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Guia'
guias = {guias_repr}
for i, g in enumerate(guias, 2):
    ws['A'+str(i)] = g
wb.save('/data/GUIAS.xlsx')
print('OK')
" 2>&1''')
        print("   ", out2.strip())
        out, _ = run_ssh(client, "ls -la /home/pdiaz/WEBSITE/")
        print("   Contenido WEBSITE:", out)

        # 4b. Subir script actualizado para montar en contenedor
        script_local = os.path.join(os.path.dirname(__file__) or ".", "test_selenium_browser.py")
        if os.path.exists(script_local):
            try:
                sftp = client.open_sftp()
                sftp.put(script_local, "/home/pdiaz/test_selenium_browser.py")
                sftp.close()
                print("   Script actualizado subido para montar en contenedor")
            except Exception as e:
                print(f"   [WARNING] No se pudo subir script: {e}")

        # 4c. Subir manuales de usuario y técnico
        run_ssh(client, "mkdir -p /home/pdiaz/docs")
        try:
            sftp = client.open_sftp()
            for manual in ["MANUAL_USUARIO.md", "MANUAL_TECNICO.md"]:
                manual_local = os.path.join(os.path.dirname(__file__) or ".", manual)
                if os.path.exists(manual_local):
                    sftp.put(manual_local, f"/home/pdiaz/docs/{manual}")
                    print(f"   Manual {manual} subido")
            sftp.close()
        except Exception as e:
            print(f"   [WARNING] No se pudieron subir manuales: {e}")

        # 5. Ejecutar el contenedor (modo detached, con script actualizado si está disponible)
        print("\n=== 3. EJECUTANDO CONTENEDOR padua-pod ===")
        print("\n   Para VER la ejecucion en vivo, abre otra terminal y ejecuta:")
        print(f"   ssh {USER}@{HOST}")
        print("   docker logs -f padua-pod-run")
        print("   (Ctrl+C para dejar de seguir los logs)\n")
        run_cmd = """docker run -d --name padua-pod-run -e RUNNING_IN_DOCKER=true \
  -v /home/pdiaz/WEBSITE:/app/WEBSITE \
  -v /home/pdiaz/DESCARGA:/app/DESCARGA \
  -v /home/pdiaz/docs:/app/docs:ro \
  -v /home/pdiaz/test_selenium_browser.py:/app/test_selenium_browser.py:ro \
  padua-pod:latest 2>&1"""
        out, err = run_ssh(client, run_cmd)
        print(out)
        if err:
            print("stderr:", err)

        # 6. Esperar y capturar logs
        print("\n=== 4. ESPERANDO EJECUCIÓN (30 seg) ===")
        time.sleep(30)
        
        print("\n=== 5. LOGS DE EJECUCIÓN ===")
        logs, _ = run_ssh(client, "docker logs --tail 500 padua-pod-run 2>&1", timeout=30)
        logs_safe = logs.replace("\u26a0", "[WARN]").replace("\u2713", "[OK]").replace("\u2717", "[FAIL]")
        print(logs_safe)

        # 7. Estado final
        print("\n=== 6. ESTADO FINAL ===")
        out, _ = run_ssh(client, "docker ps -a --filter name=padua-pod-run")
        print(out)
        
        # 8. Validar errores en logs
        if logs:
            logs_lower = logs.lower()
            errors = []
            if "fail" in logs_lower or "[fail]" in logs_lower:
                errors.append("- Mensajes [FAIL] encontrados")
            if "error" in logs_lower or "exception" in logs_lower:
                errors.append("- Errores o excepciones en logs")
            if "nenvio" in logs_lower and "no encontrado" in logs_lower:
                errors.append("- Campo nenvio no encontrado (problema de frame/contexto)")
            if "timeout" in logs_lower:
                errors.append("- Timeouts detectados")
            if errors:
                print("\n[VALIDACIÓN] Posibles errores:")
                for e in errors:
                    print(e)
            else:
                print("\n[OK] No se detectaron errores obvios en la validación automática.")

        client.close()
        print("\n[OK] Proceso completado.")
    except Exception as e:
        print(f"[FAIL] Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
