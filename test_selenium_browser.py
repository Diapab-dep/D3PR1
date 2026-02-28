from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
import os
import sys
import time
from datetime import datetime
import urllib.request
import urllib.parse
import openpyxl
import shutil
import re
try:
    import requests
    requests_available = True
except ImportError:
    requests_available = False

# Intentar importar pyautogui para escribir directamente donde está el cursor
try:
    import pyautogui
    pyautogui_available = True
except ImportError:
    pyautogui_available = False
    print("[WARNING] pyautogui no esta instalado. Instalalo con: pip install pyautogui")

def is_docker():
    """Detecta si el script está corriendo dentro de un contenedor Docker."""
    return os.environ.get("RUNNING_IN_DOCKER", "").lower() == "true"


def safe_input(prompt=""):
    """Input seguro que no falla en entornos no-interactivos (Docker, cron)."""
    if is_docker():
        print(f"{prompt} [Omitido en modo Docker]")
        return ""
    try:
        return input(prompt)
    except (EOFError, OSError):
        print(f"{prompt} [Omitido - entorno no interactivo]")
        return ""


def main():
    driver = None
    try:
        # === RUTAS BASE DEL WORKSPACE ===
        WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
        WEBSITE_DIR = os.path.join(WORKSPACE_DIR, "WEBSITE")
        DESCARGA_DIR = os.path.join(WORKSPACE_DIR, "DESCARGA")
        os.makedirs(WEBSITE_DIR, exist_ok=True)
        os.makedirs(DESCARGA_DIR, exist_ok=True)
        print(f"Workspace: {WORKSPACE_DIR}")
        print(f"Carpeta WEBSITE: {WEBSITE_DIR}")
        print(f"Carpeta DESCARGA: {DESCARGA_DIR}")
        if is_docker():
            print("🐳 Ejecutando en modo Docker")

        print("=" * 60)
        print("PRUEBA DE APERTURA DE NAVEGADOR - PADUA")
        print("=" * 60)
        print("\nIniciando Chrome...")
        
        # Configurar opciones de Chrome
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")  # Maximizar ventana
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--ignore-ssl-errors")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--allow-running-insecure-content")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        # Optimizaciones de rendimiento (sin duplicados)
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-background-networking")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-ipc-flooding-protection")
        chrome_options.add_argument("--disable-hang-monitor")
        chrome_options.add_argument("--disable-prompt-on-repost")
        chrome_options.add_argument("--disable-domain-reliability")
        chrome_options.add_argument("--disable-component-update")
        chrome_options.add_argument("--disable-sync")
        chrome_options.add_argument("--disable-translate")
        chrome_options.add_argument("--disable-default-apps")
        chrome_options.add_argument("--no-first-run")
        chrome_options.add_argument("--no-default-browser-check")
        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-password-generation")
        chrome_options.add_argument("--disable-save-password-bubble")
        chrome_options.add_argument("--disable-session-crashed-bubble")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-breakpad")
        chrome_options.add_argument("--disable-client-side-phishing-detection")
        chrome_options.add_argument("--disable-component-extensions-with-background-pages")
        chrome_options.add_argument("--disable-extensions-file-access-check")
        chrome_options.add_argument("--disable-extensions-http-throttling")
        chrome_options.add_argument("--disable-print-preview")
        chrome_options.add_argument("--disable-speech-api")
        chrome_options.add_argument("--hide-scrollbars")
        chrome_options.add_argument("--mute-audio")
        chrome_options.add_argument("--metrics-recording-only")
        chrome_options.add_argument("--safebrowsing-disable-auto-update")
        chrome_options.add_argument("--enable-automation")
        chrome_options.add_argument("--password-store=basic")
        chrome_options.add_argument("--use-mock-keychain")
        # Deshabilitar características innecesarias para mejorar rendimiento
        chrome_options.add_argument("--disable-features=IsolateOrigins,site-per-process,TranslateUI,AudioServiceOutOfProcess,MediaRouter,DialMediaRouteProvider,CastMediaRouteProvider,HardwareMediaKeyHandling,GlobalMediaControls,LiveCaption,OptimizationHints,CalculateNativeWinOcclusion,RendererScheduling,ThrottleForegroundTimers,ThrottleBackgroundTimers,BlockInsecurePrivateNetworkRequests,CertificateTransparencyComponentUpdater,NetworkService,NetworkServiceInProcess,NetworkServiceLogging,OutOfBlinkCors,RendererCodeIntegrity")

        # === OPCIONES ADICIONALES PARA DOCKER / LINUX ===
        if is_docker():
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--remote-debugging-port=9222")
            # NO usamos --headless porque necesitamos Xvfb + PyAutoGUI
            print("  [Docker] Opciones Chrome: --no-sandbox, --disable-dev-shm-usage, --disable-gpu")

        # Configurar carpeta de descargas de Chrome dentro del workspace
        prefs = {
            "download.default_directory": DESCARGA_DIR,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        print("Configurando ChromeDriver...")
        
        # === MÉTODO DOCKER: ChromeDriver ya instalado en /usr/local/bin ===
        if is_docker():
            print("  [Docker] Usando ChromeDriver del sistema (/usr/local/bin/chromedriver)...")
            service = Service("/usr/local/bin/chromedriver")
            driver = webdriver.Chrome(service=service, options=chrome_options)
            print("[OK] Chrome iniciado correctamente en modo Docker.")

        # === MÉTODO LOCAL (Windows): chromedriver-autoinstaller o búsqueda manual ===
        else:
            # Método 1: Intentar con chromedriver-autoinstaller (más simple)
            try:
                import chromedriver_autoinstaller
                print("Usando chromedriver-autoinstaller...")
                chromedriver_autoinstaller.install()
                driver = webdriver.Chrome(options=chrome_options)
                print("[OK] Chrome iniciado correctamente con chromedriver-autoinstaller.")
                
            except ImportError:
                print("chromedriver-autoinstaller no está instalado.")
                print("Intentando método alternativo...")
                
                # Método 2: Buscar ChromeDriver en ubicaciones comunes
                script_dir = os.path.dirname(os.path.abspath(__file__))
                possible_paths = [
                    os.path.join(script_dir, "chromedriver.exe"),
                    os.path.join(os.getcwd(), "chromedriver.exe"),
                    r"C:\chromedriver\chromedriver.exe",
                    r"C:\Program Files\chromedriver\chromedriver.exe",
                    r"C:\Program Files (x86)\chromedriver\chromedriver.exe",
                ]
                
                driver_path = None
                for path in possible_paths:
                    if os.path.exists(path):
                        driver_path = path
                        print(f"[OK] ChromeDriver encontrado en: {driver_path}")
                        break
                
                if driver_path:
                    service = Service(driver_path)
                    driver = webdriver.Chrome(service=service, options=chrome_options)
                    print("[OK] Chrome iniciado correctamente.")
                else:
                    print("\n" + "=" * 60)
                    print("ChromeDriver no encontrado.")
                    print("=" * 60)
                    print("\nSOLUCIÓN RÁPIDA:")
                    print("1. Instala chromedriver-autoinstaller (recomendado):")
                    print("   python -m pip install chromedriver-autoinstaller")
                    print("\n2. O descarga ChromeDriver manualmente:")
                    print("   - Ve a: https://googlechromelabs.github.io/chrome-for-testing/")
                    print("   - Descarga la versión compatible con tu Chrome")
                    print("   - Extrae chromedriver.exe")
                    print("   - Colócalo en esta carpeta:", script_dir)
                    print("\n" + "=" * 60)
                    safe_input("\nPulsa ENTER para salir...")
                    return
                    
            except Exception as e:
                print(f"\n[FAIL] Error al iniciar Chrome: {e}")
                print("\n" + "=" * 60)
                print("SOLUCIÓN:")
                print("=" * 60)
                print("\nInstala chromedriver-autoinstaller ejecutando:")
                print("python -m pip install chromedriver-autoinstaller")
                print("\nLuego ejecuta este script de nuevo.")
                safe_input("\nPulsa ENTER para salir...")
                return
        
        # Si llegamos aquí, Chrome está abierto
        print("\n" + "=" * 60)
        print("Navegador abierto. Cargando URL de Padua...")
        print("=" * 60)

        # Cargar la URL principal
        driver.get("https://alertran.latinlogistics.com.co/padua/inicio.do")
        print("[OK] URL cargada correctamente.")

        # Esperar a que la página cargue completamente
        wait = WebDriverWait(driver, 10)
        # Reducido de 2s a 0.5s - WebDriverWait ya maneja la espera
        time.sleep(0.5)

        print("\n" + "=" * 60)
        print("INICIANDO PROCESO DE LOGIN")
        print("=" * 60)

        # Credenciales
        usuario = "A023863399"
        contraseña = "*0P52ryr74g!"

        print(f"\nUsuario: {usuario}")
        print("Buscando campos de login...")

        # Buscar campo de usuario (intentar diferentes selectores)
        campo_usuario = None
        selectores_usuario = [
            (By.NAME, "usuario"),
            (By.ID, "usuario"),
            (By.NAME, "username"),
            (By.ID, "username"),
            (By.XPATH, "//input[@type='text']"),
            (By.XPATH, "//input[contains(@name, 'usuario') or contains(@id, 'usuario')]"),
        ]

        for by, selector in selectores_usuario:
            try:
                campo_usuario = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"[OK] Campo de usuario encontrado: {by} = '{selector}'")
                break
            except TimeoutException:
                continue

        if not campo_usuario:
            # Intentar encontrar cualquier input de texto
            try:
                inputs = driver.find_elements(By.TAG_NAME, "input")
                for inp in inputs:
                    if inp.get_attribute("type") in ["text", None, ""]:
                        campo_usuario = inp
                        print("[OK] Campo de usuario encontrado (método alternativo)")
                        break
            except Exception:
                pass

        if not campo_usuario:
            raise Exception("No se pudo encontrar el campo de usuario")

        # Limpiar y escribir usuario
        campo_usuario.clear()
        campo_usuario.send_keys(usuario)
        print(f"[OK] Usuario ingresado: {usuario}")
        time.sleep(0.5)

        # Buscar campo de contraseña
        campo_contraseña = None
        selectores_contraseña = [
            (By.NAME, "contraseña"),
            (By.NAME, "password"),
            (By.ID, "contraseña"),
            (By.ID, "password"),
            (By.XPATH, "//input[@type='password']"),
            (By.XPATH, "//input[contains(@name, 'contraseña') or contains(@name, 'password')]"),
        ]

        for by, selector in selectores_contraseña:
            try:
                campo_contraseña = driver.find_element(by, selector)
                print(f"[OK] Campo de contraseña encontrado: {by} = '{selector}'")
                break
            except NoSuchElementException:
                continue

        if not campo_contraseña:
            # Buscar input de tipo password
            try:
                campo_contraseña = driver.find_element(By.XPATH, "//input[@type='password']")
                print("[OK] Campo de contraseña encontrado (método alternativo)")
            except Exception:
                raise Exception("No se pudo encontrar el campo de contraseña")

        # Limpiar y escribir contraseña
        campo_contraseña.clear()
        campo_contraseña.send_keys(contraseña)
        print("[OK] Contraseña ingresada")
        time.sleep(0.5)

        # Enviar el formulario de forma robusta
        print("\nBuscando formulario de login para enviarlo...")
        form = None
        try:
            # Tomar el formulario padre del campo de contraseña
            form = campo_contraseña.find_element(By.XPATH, "ancestor::form[1]")
            print("[OK] Formulario de login encontrado como ancestro del campo de contraseña.")
        except Exception:
            try:
                # Como alternativa, tomar el primer <form> de la página
                form = driver.find_element(By.TAG_NAME, "form")
                print("[OK] Formulario de login encontrado como primer <form> de la página.")
            except Exception:
                form = None

        if form is not None:
            try:
                print("Enviando formulario con JavaScript (form.submit())...")
                driver.execute_script("arguments[0].submit();", form)
                print("[OK] Formulario enviado correctamente.")
            except Exception as e:
                print(f"Error al enviar el formulario con JS: {e}")
                print("Intentando enviar ENTER en el campo de contraseña...")
                campo_contraseña.send_keys(Keys.ENTER)
                print("[OK] ENTER enviado en el campo de contraseña.")
        else:
            print("No se pudo localizar el formulario. Enviando ENTER en el campo de contraseña...")
            campo_contraseña.send_keys(Keys.ENTER)
            print("[OK] ENTER enviado en el campo de contraseña.")

        # Esperar a que la página cambie o cargue (optimizado)
        print("\nEsperando respuesta del servidor...")
        time.sleep(1)  # Reducido de 3s a 1s

        # Verificar si el login fue exitoso o si hay mensajes de error
        current_url = driver.current_url
        page_source = driver.page_source.lower()

        print("\n" + "=" * 60)
        print("RESULTADO DEL LOGIN")
        print("=" * 60)
        print(f"URL actual: {current_url}")

        # Verificar mensajes de error comunes
        errores_posibles = [
            "usuario incorrecto",
            "contraseña incorrecta",
            "usuario bloqueado",
            "usuario en proceso de actualización",
            "validando usuario"
        ]

        error_encontrado = None
        for error in errores_posibles:
            if error in page_source:
                error_encontrado = error
                print(f"\n[WARNING] Mensaje detectado: '{error}'")
                break

        if not error_encontrado:
            # Verificar si el login fue exitoso de múltiples formas
            login_exitoso = False
            
            # Método 1: Verificar si la URL cambió
            if "inicio.do" not in current_url or "menu" in current_url.lower() or "dashboard" in current_url.lower():
                login_exitoso = True
                print("\n[OK] Login exitoso detectado - La URL cambió")
            
            # Método 2: Verificar si aparecen elementos típicos de la página principal (menús, botones)
            if not login_exitoso:
                try:
                    # Buscar elementos que aparecen después del login
                    elementos_post_login = [
                        (By.CLASS_NAME, "pui-button-text"),  # Botones del menú
                        (By.CLASS_NAME, "ui-menuitem-text"),  # Items del menú
                        (By.XPATH, "//*[contains(@class,'menu')]"),
                        (By.XPATH, "//*[contains(@class,'nav')]"),
                        (By.XPATH, "//*[contains(@class,'sidebar')]"),
                    ]
                    
                    for by, selector in elementos_post_login:
                        try:
                            elementos = driver.find_elements(by, selector)
                            if len(elementos) > 0:
                                login_exitoso = True
                                print(f"\n[OK] Login exitoso detectado - Se encontraron elementos de la página principal ({by})")
                                break
                        except:
                            continue
                except:
                    pass
            
            # Método 3: Verificar si desaparecieron los campos de login
            if not login_exitoso:
                try:
                    campos_login = driver.find_elements(By.XPATH, "//input[@type='password']")
                    if len(campos_login) == 0:
                        login_exitoso = True
                        print("\n[OK] Login exitoso detectado - Los campos de login desaparecieron")
                except:
                    pass
            
            if login_exitoso:
                # Intentar navegar al menú 7.8 usando el buscador en la parte superior derecha
                print("\n" + "=" * 60)
                print("NAVEGACIÓN A MENÚ 7.8 USANDO EL BUSCADOR")
                print("=" * 60)
                # guia_prueba se asigna desde el Excel en el loop de descarga masiva

                try:
                    # Esperar a que la página cargue completamente después del login
                    print("Esperando a que la página principal cargue completamente...")
                    time.sleep(5)
                    
                    # CAPTURAR la URL principal después del login (para recuperación)
                    url_principal_app = driver.current_url
                    print(f"\nURL principal capturada: {url_principal_app}")
                    
                    # Verificar el estado de la página
                    print(f"URL actual después del login: {driver.current_url}")
                    print(f"Título de la página: {driver.title}")
                    
                    # Verificar si hay framesets o iframes
                    print("\nVerificando estructura de frames de la página...")
                    
                    # Intentar cambiar al frame "menu" (según el HTML proporcionado)
                    print("   Intentando cambiar al frame 'menu'...")
                    try:
                        driver.switch_to.frame("menu")
                        print("[OK] Cambiado al frame 'menu'")
                        time.sleep(2)  # Esperar a que el frame cargue
                    except Exception as e:
                        print(f"   No se pudo cambiar al frame 'menu' por nombre: {e}")
                        # Intentar buscar frames por tag
                        try:
                            frames = driver.find_elements(By.TAG_NAME, "frame")
                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                            print(f"   Encontrados {len(frames)} frames y {len(iframes)} iframes")
                            
                            # Intentar cambiar al primer frame
                            if len(frames) > 0:
                                for idx, frame in enumerate(frames):
                                    try:
                                        frame_name = frame.get_attribute("name") or ""
                                        frame_src = frame.get_attribute("src") or ""
                                        print(f"   Frame #{idx+1}: name='{frame_name}', src='{frame_src[:50]}...'")
                                        if frame_name == "menu" or "menu" in frame_src.lower():
                                            driver.switch_to.frame(frame)
                                            print(f"[OK] Cambiado al frame 'menu' (índice {idx+1})")
                                            time.sleep(2)
                                            break
                                    except Exception as e2:
                                        print(f"   Error al cambiar al frame #{idx+1}: {e2}")
                            
                            # Si no encontró por frames, intentar iframes
                            if driver.current_window_handle == driver.window_handles[0]:
                                if len(iframes) > 0:
                                    for idx, iframe in enumerate(iframes):
                                        try:
                                            iframe_name = iframe.get_attribute("name") or ""
                                            iframe_src = iframe.get_attribute("src") or ""
                                            print(f"   Iframe #{idx+1}: name='{iframe_name}', src='{iframe_src[:50]}...'")
                                            if "menu" in iframe_name.lower() or "menu" in iframe_src.lower():
                                                driver.switch_to.frame(iframe)
                                                print(f"[OK] Cambiado al iframe 'menu' (índice {idx+1})")
                                                time.sleep(2)
                                                break
                                        except Exception as e3:
                                            print(f"   Error al cambiar al iframe #{idx+1}: {e3}")
                        except Exception as e4:
                            print(f"   Error al buscar frames: {e4}")
                    
                    # Ir al inicio del frame
                    try:
                        driver.execute_script("window.scrollTo(0, 0);")
                        time.sleep(1)
                    except:
                        pass
                    
                    # PASO 1: Buscar el campo de búsqueda en la parte superior derecha (cerca de "ALERTRAN - Padua")
                    print("\n[PASO 1] Buscando campo de búsqueda en el frame actual...")
                    campo_buscador = None
                    
                    # Intentar buscar el campo directamente con WebDriverWait (más robusto)
                    try:
                        print("   Intentando encontrar campo por name='funcionalidad_codigo' con espera...")
                        campo_buscador = WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.NAME, "funcionalidad_codigo"))
                        )
                        if campo_buscador.is_displayed():
                            print("[OK] Campo de búsqueda encontrado por name='funcionalidad_codigo'")
                    except TimeoutException:
                        print("   Campo no encontrado por name, intentando otras estrategias...")
                        campo_buscador = None
                    except Exception as e:
                        print(f"   Error al buscar campo: {e}")
                        campo_buscador = None
                    
                    # Estrategias priorizadas: primero buscar el campo específico por name y class
                    selectores_buscador = [
                        # Prioridad 1: Campo específico por name (más confiable)
                        (By.NAME, "funcionalidad_codigo"),
                        # Prioridad 2: Campo específico por class
                        (By.XPATH, "//input[@name='funcionalidad_codigo']"),
                        (By.XPATH, "//input[contains(@class,'codigo') and contains(@class,'pui-inputtext')]"),
                        (By.XPATH, "//input[@class='codigo pui-inputtext ui-widget ui-state-default ui-corner-all']"),
                        (By.XPATH, "//input[contains(@class,'codigo')]"),
                        # Prioridad 3: Input inmediatamente después del texto "ALERTRAN - Padua 26.01.16"
                        (By.XPATH, "//*[contains(text(),'ALERTRAN - Padua')]/following::input[1]"),
                        (By.XPATH, "//*[contains(text(),'ALERTRAN - Padua 26.01.16')]/following::input[1]"),
                        (By.XPATH, "//*[contains(text(),'ALERTRAN') and contains(text(),'Padua')]/following::input[1]"),
                        # Prioridad 4: Inputs en la parte superior de la página (Y < 200)
                        (By.XPATH, "//input[@type='search']"),
                        (By.XPATH, "//input[@type='text' and (contains(@placeholder,'buscar') or contains(@placeholder,'search'))]"),
                        (By.XPATH, "//input[contains(@class,'search') or contains(@id,'search') or contains(@name,'search')]"),
                        (By.XPATH, "//input[contains(@placeholder,'Buscar')]"),
                        (By.XPATH, "//input[@placeholder]"),  # Cualquier input con placeholder
                        (By.XPATH, "//input[@type='text']"),  # Cualquier input de texto
                    ]
                    
                    for by, selector in selectores_buscador:
                        try:
                            # Para selectores por NAME, usar find_element directamente (más rápido)
                            if by == By.NAME:
                                try:
                                    campo_buscador = wait.until(EC.presence_of_element_located((by, selector)))
                                    if campo_buscador.is_displayed():
                                        print(f"[OK] Campo de búsqueda encontrado: {by} = '{selector}'")
                                        break
                                except:
                                    continue
                            else:
                                elementos = driver.find_elements(by, selector)
                                for elem in elementos:
                                    # Verificar que esté visible
                                    try:
                                        if elem.is_displayed():
                                            # Para selectores específicos (funcionalidad_codigo, codigo), aceptar directamente
                                            if "funcionalidad_codigo" in str(selector) or "codigo" in str(selector):
                                                campo_buscador = elem
                                                print(f"[OK] Campo de búsqueda encontrado: {by} = '{selector}'")
                                                break
                                            # Para selectores basados en ALERTRAN, aceptar directamente
                                            if "ALERTRAN" in str(selector) or "Padua" in str(selector):
                                                campo_buscador = elem
                                                print(f"[OK] Campo de búsqueda encontrado: {by} = '{selector}'")
                                                break
                                            # Para otros selectores, verificar posición (debe estar en la parte superior)
                                            location = elem.location
                                            if location['y'] < 200:  # En la parte superior de la página
                                                campo_buscador = elem
                                                print(f"[OK] Campo de búsqueda encontrado: {by} = '{selector}' (posición Y: {location['y']})")
                                                break
                                    except:
                                        continue
                                if campo_buscador:
                                    break
                        except:
                            continue
                    
                    if not campo_buscador:
                        # Método alternativo: Buscar todos los inputs visibles en la parte superior
                        try:
                            print("Buscando todos los inputs visibles en la parte superior de la página...")
                            todos_inputs = driver.find_elements(By.XPATH, "//input")
                            print(f"   Encontrados {len(todos_inputs)} inputs en total")
                            
                            candidatos = []
                            for inp in todos_inputs:
                                try:
                                    if inp.is_displayed():
                                        location = inp.location
                                        size = inp.size
                                        tipo = inp.get_attribute("type") or "text"
                                        inp_id = inp.get_attribute("id") or ""
                                        inp_name = inp.get_attribute("name") or ""
                                        inp_class = inp.get_attribute("class") or ""
                                        
                                        # Mostrar información de inputs en la parte superior
                                        if location['y'] < 300:  # Parte superior de la página
                                            info = f"tipo='{tipo}', Y={location['y']}, X={location['x']}, tamaño={size['width']}x{size['height']}"
                                            if inp_id:
                                                info += f", id='{inp_id}'"
                                            if inp_name:
                                                info += f", name='{inp_name}'"
                                            if inp_class:
                                                info += f", class='{inp_class[:30]}'"
                                            print(f"   Input encontrado: {info}")
                                            
                                            # Priorizar inputs pequeños en la parte superior (como el buscador)
                                            if location['y'] < 200 and size['width'] < 300 and size['height'] < 50:
                                                if tipo in ["text", "search", None, ""]:
                                                    candidatos.append((inp, location['y'], location['x']))
                                except:
                                    continue
                                                    
                            # Ordenar candidatos por posición Y (más arriba primero) y luego X (más a la izquierda)
                            if candidatos:
                                candidatos.sort(key=lambda x: (x[1], x[2]))
                                campo_buscador = candidatos[0][0]
                                print(f"[OK] Campo de búsqueda encontrado (input pequeño en parte superior, Y={candidatos[0][1]}, X={candidatos[0][2]})")
                            
                            if not campo_buscador:
                                # Buscar otros tipos de elementos que puedan ser el buscador (divs editables, textareas, etc.)
                                print("Buscando otros tipos de elementos editables (divs, textareas, etc.)...")
                                try:
                                    # Buscar divs con contenteditable
                                    divs_editables = driver.find_elements(By.XPATH, "//div[@contenteditable='true'] | //div[@contenteditable]")
                                    print(f"   Encontrados {len(divs_editables)} divs editables")
                                    for div in divs_editables:
                                        if div.is_displayed():
                                            loc = div.location
                                            if loc['y'] < 200:
                                                print(f"   Div editable encontrado: Y={loc['y']}, X={loc['x']}, id='{div.get_attribute('id')}', class='{div.get_attribute('class')[:50]}'")
                                                if loc['y'] < 200 and div.size['width'] < 300:
                                                    campo_buscador = div
                                                    print(f"[OK] Campo de búsqueda encontrado (div editable)")
                                                    break
                                    
                                    # Buscar textareas
                                    if not campo_buscador:
                                        textareas = driver.find_elements(By.XPATH, "//textarea")
                                        for ta in textareas:
                                            if ta.is_displayed():
                                                loc = ta.location
                                                if loc['y'] < 200:
                                                    campo_buscador = ta
                                                    print(f"[OK] Campo de búsqueda encontrado (textarea)")
                                                    break
                                except Exception as e:
                                    print(f"   Error al buscar elementos editables: {e}")
                            
                            if not campo_buscador:
                                # Primero, buscar TODOS los elementos interactivos en la parte superior
                                print("Buscando TODOS los elementos interactivos en la parte superior de la página...")
                                try:
                                    script_todos = """
                                    var elementos = document.querySelectorAll('input, textarea, [contenteditable="true"], [contenteditable], [role="textbox"], div[onclick], span[onclick]');
                                    var resultados = [];
                                    for (var i = 0; i < elementos.length; i++) {
                                        if (elementos[i].offsetParent !== null) {
                                            var rect = elementos[i].getBoundingClientRect();
                                            if (rect.top < 200 && rect.width < 500 && rect.height < 100) {
                                                resultados.push({
                                                    tagName: elementos[i].tagName,
                                                    id: elementos[i].id || '',
                                                    className: elementos[i].className || '',
                                                    top: rect.top,
                                                    left: rect.left,
                                                    width: rect.width,
                                                    height: rect.height
                                                });
                                            }
                                        }
                                    }
                                    return resultados;
                                    """
                                    todos_interactivos = driver.execute_script(script_todos)
                                    print(f"   Encontrados {len(todos_interactivos)} elementos interactivos en la parte superior")
                                    for elem_info in todos_interactivos:
                                        print(f"      - {elem_info['tagName']}: id='{elem_info['id']}', class='{elem_info['className'][:40]}', pos=({elem_info['left']}, {elem_info['top']}), tamaño={elem_info['width']}x{elem_info['height']}")
                                        
                                        # Si es un elemento pequeño en la parte superior, podría ser el buscador
                                        if elem_info['width'] < 300 and elem_info['height'] < 50 and elem_info['top'] < 150:
                                            try:
                                                if elem_info['id']:
                                                    campo_buscador = driver.find_element(By.ID, elem_info['id'])
                                                    print(f"[OK] Campo de búsqueda encontrado (elemento pequeño en parte superior, ID: {elem_info['id']})")
                                                    break
                                                elif elem_info['className']:
                                                    primera_clase = elem_info['className'].split()[0]
                                                    if primera_clase:
                                                        campo_buscador = driver.find_element(By.CLASS_NAME, primera_clase)
                                                        print(f"[OK] Campo de búsqueda encontrado (elemento pequeño en parte superior, Class: {primera_clase})")
                                                        break
                                            except:
                                                pass
                                except Exception as e:
                                    print(f"   Error al buscar todos los elementos interactivos: {e}")
                                
                                # Buscar cualquier elemento cerca del texto "ALERTRAN" o "Padua" usando JavaScript
                                if not campo_buscador:
                                    print("Buscando campo de búsqueda cerca del texto 'ALERTRAN' o 'Padua' usando JavaScript...")
                                    elementos_texto = driver.find_elements(By.XPATH, "//*[contains(text(),'ALERTRAN') or contains(text(),'Padua')]")
                                    print(f"   Encontrados {len(elementos_texto)} elementos con texto 'ALERTRAN' o 'Padua'")
                                    
                                    for idx, elem_texto in enumerate(elementos_texto):
                                        try:
                                            # Obtener información del elemento
                                            tag = elem_texto.tag_name
                                            elem_id = elem_texto.get_attribute("id") or ""
                                            elem_class = elem_texto.get_attribute("class") or ""
                                            
                                            # Intentar obtener texto y ubicación
                                            try:
                                                texto_elem = elem_texto.text.strip()
                                            except:
                                                texto_elem = elem_texto.get_attribute("textContent") or elem_texto.get_attribute("innerText") or ""
                                            
                                            try:
                                                loc = elem_texto.location
                                                loc_str = f"Y={loc['y']}, X={loc['x']}"
                                            except:
                                                loc_str = "ubicación no disponible"
                                            
                                            print(f"   Elemento ALERTRAN #{idx+1}: tag='{tag}', id='{elem_id}', class='{elem_class[:50]}', {loc_str}, texto='{texto_elem[:50]}...'")
                                            
                                            if not ("ALERTRAN" in texto_elem or "Padua" in texto_elem):
                                                continue
                                            
                                            # Usar JavaScript para buscar elementos interactivos cerca
                                            try:
                                                script = """
                                                var elem = arguments[0];
                                                var rect1 = elem.getBoundingClientRect();
                                                var elementos = document.querySelectorAll('input, textarea, [contenteditable="true"], [contenteditable], [role="textbox"]');
                                                var minDist = Infinity;
                                                var closestElem = null;
                                                
                                                for (var i = 0; i < elementos.length; i++) {
                                                    if (elementos[i].offsetParent !== null) {
                                                        var rect2 = elementos[i].getBoundingClientRect();
                                                        // Calcular distancia horizontal (a la derecha del texto ALERTRAN)
                                                        var distX = rect2.left - rect1.right;
                                                        var distY = Math.abs(rect2.top - rect1.top);
                                                        var dist = distX + distY;
                                                        
                                                        // Debe estar a la derecha y aproximadamente a la misma altura
                                                        if (distX > 0 && distX < 500 && distY < 50 && rect2.top < 200) {
                                                            if (dist < minDist) {
                                                                minDist = dist;
                                                                closestElem = elementos[i];
                                                            }
                                                        }
                                                    }
                                                }
                                                
                                                if (closestElem) {
                                                    return {
                                                        id: closestElem.id || '',
                                                        className: closestElem.className || '',
                                                        tagName: closestElem.tagName || '',
                                                        outerHTML: closestElem.outerHTML.substring(0, 200)
                                                    };
                                                }
                                                return null;
                                                """
                                                resultado_js = driver.execute_script(script, elem_texto)
                                                if resultado_js:
                                                    print(f"   Elemento cercano encontrado: tag='{resultado_js.get('tagName', '')}', id='{resultado_js.get('id', '')}', class='{resultado_js.get('className', '')[:50]}'")
                                                    
                                                    # Intentar encontrar el elemento usando los atributos
                                                    if resultado_js.get('id'):
                                                        try:
                                                            campo_buscador = driver.find_element(By.ID, resultado_js['id'])
                                                            print(f"[OK] Campo de búsqueda encontrado usando ID: {resultado_js['id']}")
                                                            break
                                                        except:
                                                            pass
                                                    
                                                    # Si no funciona por ID, intentar por clase
                                                    if not campo_buscador and resultado_js.get('className'):
                                                        try:
                                                            primera_clase = resultado_js['className'].split()[0] if resultado_js['className'] else None
                                                            if primera_clase:
                                                                campo_buscador = driver.find_element(By.CLASS_NAME, primera_clase)
                                                                print(f"[OK] Campo de búsqueda encontrado usando clase: {primera_clase}")
                                                                break
                                                        except:
                                                            pass
                                                    
                                                    # Si aún no funciona, buscar por XPath usando el tag
                                                    if not campo_buscador:
                                                        tag_name = resultado_js.get('tagName', '').lower()
                                                        if tag_name:
                                                            try:
                                                                if tag_name == 'input':
                                                                    campo_buscador = driver.find_element(By.XPATH, f"//input[@id='{resultado_js.get('id', '')}']" if resultado_js.get('id') else f"//input[contains(@class,'{resultado_js.get('className', '').split()[0]}')]")
                                                                elif tag_name == 'div':
                                                                    campo_buscador = driver.find_element(By.XPATH, f"//div[@id='{resultado_js.get('id', '')}']" if resultado_js.get('id') else f"//div[contains(@class,'{resultado_js.get('className', '').split()[0]}')]")
                                                                if campo_buscador:
                                                                    print(f"[OK] Campo de búsqueda encontrado usando XPath con tag {tag_name}")
                                                                    break
                                                            except:
                                                                pass
                                            except Exception as js_error:
                                                print(f"   Error en JavaScript: {js_error}")
                                            
                                            # Fallback: buscar elementos cercanos con XPath
                                            if not campo_buscador:
                                                try:
                                                    elementos_cercanos = elem_texto.find_elements(By.XPATH, ".//following::input[1] | .//following-sibling::input[1] | .//following::div[@contenteditable][1] | .//following-sibling::div[@contenteditable][1]")
                                                    for elem_cercano in elementos_cercanos:
                                                        if elem_cercano.is_displayed():
                                                            loc_cercano = elem_cercano.location
                                                            if loc_cercano['y'] < 200:
                                                                campo_buscador = elem_cercano
                                                                print(f"[OK] Campo de búsqueda encontrado cerca de 'ALERTRAN/Padua' (XPath)")
                                                                break
                                                except:
                                                    pass
                                            
                                            if campo_buscador:
                                                break
                                        except Exception as e:
                                            print(f"   Error al procesar elemento: {e}")
                                            continue
                        except Exception as e:
                            print(f"Error al buscar campo cerca de ALERTRAN: {e}")
                            import traceback
                            traceback.print_exc()
                    
                    if campo_buscador:
                        try:
                            # Hacer scroll al campo de búsqueda
                            driver.execute_script("arguments[0].scrollIntoView(true);", campo_buscador)
                            time.sleep(0.5)
                            
                            # Determinar el tipo de elemento y escribir "7.8" apropiadamente
                            tag_name = campo_buscador.tag_name.lower()
                            is_editable_div = campo_buscador.get_attribute("contenteditable") == "true"
                            
                            if tag_name == "div" or is_editable_div:
                                # Para divs editables, usar JavaScript
                                print("   Campo es un div editable, usando JavaScript para escribir...")
                                driver.execute_script("arguments[0].textContent = ''; arguments[0].innerText = '';", campo_buscador)
                                driver.execute_script("arguments[0].textContent = '7.8'; arguments[0].innerText = '7.8';", campo_buscador)
                                # Disparar evento input
                                driver.execute_script("""
                                    var event = new Event('input', { bubbles: true });
                                    arguments[0].dispatchEvent(event);
                                """, campo_buscador)
                                print("[OK] Texto '7.8' ingresado en el buscador (div editable)")
                            else:
                                # Para inputs normales, usar send_keys
                                campo_buscador.clear()
                                campo_buscador.send_keys("7.8")
                                print("[OK] Texto '7.8' ingresado en el buscador")
                            time.sleep(1)
                            
                            # Después de escribir "7.8", los resultados deberían aparecer automáticamente
                            # o podemos presionar ENTER para activar la búsqueda
                            print("Presionando ENTER en el campo de búsqueda para activar la búsqueda...")
                            try:
                                campo_buscador.send_keys(Keys.ENTER)
                                time.sleep(2)
                            except:
                                # Si ENTER no funciona, intentar disparar el evento onkeydown que tiene el campo
                                try:
                                    driver.execute_script("""
                                        var campo = arguments[0];
                                        var event = new KeyboardEvent('keydown', {
                                            key: 'Enter',
                                            code: 'Enter',
                                            keyCode: 13,
                                            which: 13,
                                            bubbles: true
                                        });
                                        campo.dispatchEvent(event);
                                    """, campo_buscador)
                                    time.sleep(2)
                                except:
                                    pass
                            
                            # Esperar a que aparezcan los resultados de búsqueda
                            print("Esperando resultados de búsqueda...")
                            time.sleep(3)
                            
                            # Buscar el resultado "7.8 Consulta a Histórico" en los resultados
                            print("\n[PASO 2] Buscando resultado '7.8 Consulta a Histórico' en los resultados de búsqueda...")
                            resultado_78 = None
                            
                            # Primero, buscar TODOS los elementos que contengan "7.8" para ver qué hay disponible
                            print("   Buscando todos los elementos que contengan '7.8'...")
                            try:
                                todos_elementos_78 = driver.find_elements(By.XPATH, "//*[contains(text(),'7.8')]")
                                print(f"   Encontrados {len(todos_elementos_78)} elementos con '7.8'")
                                
                                for idx, elem in enumerate(todos_elementos_78[:10]):  # Mostrar solo los primeros 10
                                    try:
                                        if elem.is_displayed():
                                            texto = elem.text.strip()
                                            tag = elem.tag_name
                                            elem_id = elem.get_attribute("id") or ""
                                            elem_class = elem.get_attribute("class") or ""
                                            if texto:
                                                print(f"      Elemento #{idx+1}: tag='{tag}', texto='{texto[:60]}...', id='{elem_id}', class='{elem_class[:40]}'")
                                    except:
                                        pass
                            except Exception as e:
                                print(f"   Error al buscar elementos: {e}")
                            
                            # Buscar en menús desplegables o popups
                            print("   Buscando en menús desplegables o popups...")
                            try:
                                # Buscar elementos con clases comunes de menús desplegables
                                menus = driver.find_elements(By.XPATH, "//ul[@class] | //div[@class='ui-menu'] | //div[contains(@class,'menu')] | //div[contains(@class,'dropdown')]")
                                print(f"   Encontrados {len(menus)} posibles menús")
                                for menu in menus:
                                    try:
                                        if menu.is_displayed():
                                            items = menu.find_elements(By.XPATH, ".//*[contains(text(),'7.8')]")
                                            if items:
                                                print(f"   Menú encontrado con {len(items)} elementos que contienen '7.8'")
                                    except:
                                        pass
                            except:
                                pass
                            
                            # Esperar a que aparezcan los resultados con WebDriverWait
                            try:
                                # Esperar a que aparezca algún elemento con "7.8"
                                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'7.8')]"))
                                )
                                print("[OK] Resultados de búsqueda aparecieron")
                            except:
                                print("[WARNING] Los resultados pueden no haber aparecido, pero continuando...")
                            
                            selectores_resultado = [
                                # Prioridad 1: Texto exacto o muy específico
                                (By.XPATH, "//*[contains(text(),'7.8') and contains(text(),'Histórico')]"),
                                (By.XPATH, "//*[contains(text(),'7.8') and contains(text(),'historico')]"),
                                (By.XPATH, "//*[contains(text(),'7.8 Consulta a Histórico')]"),
                                (By.XPATH, "//*[normalize-space(text())='7.8 Consulta a Histórico']"),
                                # Prioridad 2: Enlaces y elementos de lista
                                (By.XPATH, "//a[contains(text(),'7.8') and contains(text(),'Histórico')]"),
                                (By.XPATH, "//li[contains(text(),'7.8') and contains(text(),'Histórico')]"),
                                (By.XPATH, "//span[contains(text(),'7.8') and contains(text(),'Histórico')]"),
                                (By.XPATH, "//div[contains(text(),'7.8') and contains(text(),'Histórico')]"),
                                # Prioridad 3: Cualquier elemento con "7.8" (más amplio)
                                (By.XPATH, "//a[contains(text(),'7.8')]"),
                                (By.XPATH, "//li[contains(text(),'7.8')]"),
                                (By.XPATH, "//div[contains(text(),'7.8')]"),
                                (By.XPATH, "//span[contains(text(),'7.8')]"),
                                (By.XPATH, "//td[contains(text(),'7.8')]"),
                                (By.XPATH, "//*[contains(text(),'7.8')]"),  # Cualquier elemento
                            ]
                            
                            for by, selector in selectores_resultado:
                                try:
                                    elementos = driver.find_elements(by, selector)
                                    print(f"   Probando selector: {selector} - encontrados {len(elementos)} elementos")
                                    
                                    # Si encontramos elementos con este selector, examinarlos en detalle
                                    if len(elementos) > 0 and ("Histórico" in selector or "historico" in selector.lower()):
                                        print(f"   [OK] Encontrados {len(elementos)} elementos con '7.8' y 'Histórico'")
                                        for idx, elem in enumerate(elementos):
                                            try:
                                                tag = elem.tag_name
                                                
                                                # Obtener el texto completo incluyendo hijos usando JavaScript
                                                texto_completo = driver.execute_script("""
                                                    var elem = arguments[0];
                                                    return elem.textContent || elem.innerText || '';
                                                """, elem)
                                                texto_completo = texto_completo.strip()
                                                
                                                # También obtener el texto directo
                                                texto_directo = elem.text.strip()
                                                
                                                is_displayed = elem.is_displayed()
                                                is_enabled = elem.is_enabled()
                                                
                                                print(f"      Elemento #{idx+1}: tag='{tag}', texto_directo='{texto_directo}', texto_completo='{texto_completo}', visible={is_displayed}, enabled={is_enabled}")
                                                
                                                # Usar el texto completo para verificar
                                                texto_verificar = texto_completo if texto_completo else texto_directo
                                                
                                                # Si contiene "7.8" y "Histórico", es el correcto
                                                if "7.8" in texto_verificar and ("Histórico" in texto_verificar or "historico" in texto_verificar.lower()):
                                                    # Buscar el elemento padre clickeable
                                                    elemento_clickeable = None
                                                    
                                                    # Si el elemento mismo es clickeable
                                                    if tag in ["a", "button"] or elem.get_attribute("onclick"):
                                                        elemento_clickeable = elem
                                                        print(f"         Elemento mismo es clickeable")
                                                    else:
                                                        # Buscar el elemento padre clickeable (a, li, etc.)
                                                        try:
                                                            # Buscar ancestro clickeable
                                                            padre_a = elem.find_element(By.XPATH, "./ancestor::a[1]")
                                                            if padre_a:
                                                                elemento_clickeable = padre_a
                                                                print(f"         Encontrado padre <a> clickeable")
                                                        except:
                                                            try:
                                                                padre_li = elem.find_element(By.XPATH, "./ancestor::li[1]")
                                                                if padre_li:
                                                                    elemento_clickeable = padre_li
                                                                    print(f"         Encontrado padre <li> clickeable")
                                                            except:
                                                                try:
                                                                    padre = elem.find_element(By.XPATH, "./..")
                                                                    if padre.tag_name in ["a", "li", "button"] or padre.get_attribute("onclick"):
                                                                        elemento_clickeable = padre
                                                                        print(f"         Encontrado padre '{padre.tag_name}' clickeable")
                                                                    else:
                                                                        elemento_clickeable = elem  # Usar el elemento mismo
                                                                        print(f"         Usando elemento mismo")
                                                                except:
                                                                    elemento_clickeable = elem
                                                                    print(f"         Usando elemento mismo (sin padre)")
                                                    
                                                    if elemento_clickeable:
                                                        resultado_78 = elemento_clickeable
                                                        print(f"[OK] Resultado '7.8 Consulta a Histórico' encontrado: '{texto_verificar}'")
                                                        break
                                            except Exception as e_elem:
                                                print(f"      Error al examinar elemento #{idx+1}: {e_elem}")
                                                import traceback
                                                traceback.print_exc()
                                                continue
                                    
                                    # Si no encontramos con el selector específico, buscar en todos los elementos
                                    if not resultado_78:
                                        for elem in elementos:
                                            try:
                                                # Obtener el texto completo incluyendo hijos
                                                texto_completo = driver.execute_script("""
                                                    var elem = arguments[0];
                                                    return elem.textContent || elem.innerText || '';
                                                """, elem)
                                                texto_completo = texto_completo.strip()
                                                
                                                texto_directo = elem.text.strip()
                                                texto_verificar = texto_completo if texto_completo else texto_directo
                                                
                                                # Si contiene "7.8" y "Histórico", es el correcto
                                                if "7.8" in texto_verificar and ("Histórico" in texto_verificar or "historico" in texto_verificar.lower()):
                                                    # Buscar elemento clickeable
                                                    elemento_clickeable = None
                                                    
                                                    tag = elem.tag_name
                                                    if tag in ["a", "button"] or elem.get_attribute("onclick"):
                                                        elemento_clickeable = elem
                                                    else:
                                                        try:
                                                            padre_a = elem.find_element(By.XPATH, "./ancestor::a[1]")
                                                            if padre_a:
                                                                elemento_clickeable = padre_a
                                                        except:
                                                            try:
                                                                padre_li = elem.find_element(By.XPATH, "./ancestor::li[1]")
                                                                if padre_li:
                                                                    elemento_clickeable = padre_li
                                                            except:
                                                                try:
                                                                    padre = elem.find_element(By.XPATH, "./..")
                                                                    if padre.tag_name in ["a", "li", "button"] or padre.get_attribute("onclick"):
                                                                        elemento_clickeable = padre
                                                                    else:
                                                                        elemento_clickeable = elem
                                                                except:
                                                                    elemento_clickeable = elem
                                                    
                                                    if elemento_clickeable:
                                                        resultado_78 = elemento_clickeable
                                                        print(f"[OK] Resultado '7.8 Consulta a Histórico' encontrado: '{texto_verificar}'")
                                                        break
                                            except:
                                                continue
                                    
                                    if resultado_78:
                                        break
                                except Exception as e:
                                    print(f"   Error con selector {selector}: {e}")
                                    continue
                            
                            if resultado_78:
                                try:
                                    driver.execute_script("arguments[0].scrollIntoView(true);", resultado_78)
                                    time.sleep(0.5)
                                    # Guardar las ventanas antes del clic
                                    ventanas_antes_clic = driver.window_handles
                                    
                                    resultado_78.click()
                                    print("[OK] Clic realizado en '7.8 Consulta a Histórico'")
                                    print("Esperando a que cargue la página del formulario...")
                                    time.sleep(1.5)  # Optimizado: reducido de 5s a 1.5s
                                    
                                    # Verificar si se abrió una nueva ventana (popup)
                                    ventanas_despues_clic = driver.window_handles
                                    nueva_ventana = None
                                    
                                    if len(ventanas_despues_clic) > len(ventanas_antes_clic):
                                        print(f"[OK] Nueva ventana detectada después del clic")
                                        for ventana in ventanas_despues_clic:
                                            if ventana not in ventanas_antes_clic:
                                                nueva_ventana = ventana
                                                driver.switch_to.window(ventana)
                                                print(f"[OK] Cambiado a la nueva ventana")
                                                time.sleep(3)  # Esperar a que cargue la nueva ventana
                                                break
                                    else:
                                        print("   No se detectó nueva ventana, el formulario puede estar en la misma ventana")
                                    
                                    # IMPORTANTE: Después de hacer clic, esperar a que cargue la página del formulario
                                    print("\nBuscando el formulario 'CONSULTA AL HISTÓRICO DE ENVÍOS'...")
                                    
                                    # Esperar un poco más para que cargue completamente
                                    time.sleep(3)
                                    
                                    # Intentar encontrar el elemento activo (que tiene el foco) - puede ser el campo nenvio
                                    try:
                                        elemento_activo = driver.execute_script("return document.activeElement;")
                                        if elemento_activo:
                                            tag_activo = driver.execute_script("return arguments[0].tagName.toLowerCase();", elemento_activo)
                                            name_activo = driver.execute_script("return arguments[0].name || '';", elemento_activo)
                                            if tag_activo == "input" and name_activo == "nenvio":
                                                print(f"[OK] Campo 'nenvio' tiene el foco automáticamente")
                                                # Continuar directamente al paso de escribir la guía
                                            else:
                                                print(f"   Elemento activo: tag='{tag_activo}', name='{name_activo}'")
                                    except:
                                        pass
                                    
                                    # Verificar si necesitamos cambiar de frame después del clic
                                    # El formulario puede estar en el mismo frame "menu" o en otro frame
                                    frame_encontrado = False
                                    
                                    try:
                                        # Primero, verificar si el formulario ya está en el frame actual (menu)
                                        try:
                                            WebDriverWait(driver, 5).until(
                                                EC.any_of(
                                                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'CONSULTA AL HISTÓRICO')]")),
                                                    EC.presence_of_element_located((By.NAME, "nenvio"))
                                                )
                                            )
                                            print("[OK] El formulario está en el frame actual (menu)")
                                            frame_encontrado = True
                                        except:
                                            print("   El formulario no está en el frame actual, buscando en otros frames...")
                                        
                                        # Si no está en el frame actual, buscar en otros frames
                                        if not frame_encontrado:
                                            # Volver al frame principal primero
                                            driver.switch_to.default_content()
                                            time.sleep(2)
                                            
                                            # Listar todos los frames e iframes disponibles
                                            print("   Buscando frames e iframes disponibles...")
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                            print(f"   Encontrados {len(frames)} frames y {len(iframes)} iframes")
                                            
                                            todos_frames = list(frames) + list(iframes)
                                            
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                                    frame_src = frame.get_attribute("src") or ""
                                                    frame_tag = frame.tag_name
                                                    print(f"   {frame_tag.upper()}: name='{frame_name}', id='{frame_id}', src='{frame_src[:50]}...'")
                                                except:
                                                    pass
                                            
                                            # Intentar cambiar a cada frame/iframe y buscar el formulario
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                                    frame_tag = frame.tag_name
                                                    
                                                    # Probar todos los frames, incluso "menu"
                                                    driver.switch_to.default_content()
                                                    if frame_name:
                                                        driver.switch_to.frame(frame_name)
                                                    elif frame_id:
                                                        driver.switch_to.frame(frame_id)
                                                    else:
                                                        # Si no tiene name ni id, usar el índice
                                                        indice = todos_frames.index(frame)
                                                        driver.switch_to.frame(indice)
                                                    
                                                    print(f"   Probando {frame_tag} '{frame_name or frame_id or indice}'...")
                                                    time.sleep(2)
                                                    
                                                    # Buscar el formulario en este frame
                                                    try:
                                                        WebDriverWait(driver, 8).until(  # Aumentado a 8 segundos
                                                            EC.any_of(
                                                                EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'CONSULTA AL HISTÓRICO')]")),
                                                                EC.presence_of_element_located((By.NAME, "nenvio"))
                                                            )
                                                        )
                                                        print(f"[OK] Formulario encontrado en el {frame_tag} '{frame_name or frame_id}'")
                                                        frame_encontrado = True
                                                        break
                                                    except:
                                                        print(f"   Formulario no encontrado en {frame_tag} '{frame_name or frame_id}'")
                                                except Exception as e:
                                                    print(f"   Error al cambiar al {frame_tag}: {e}")
                                                    continue
                                            
                                            # Si no se encontró en ningún frame, volver al frame "menu"
                                            if not frame_encontrado:
                                                print("[WARNING] Formulario no encontrado en ningún frame, volviendo al frame 'menu'...")
                                                try:
                                                    driver.switch_to.default_content()
                                                    driver.switch_to.frame("menu")
                                                    time.sleep(2)
                                                except:
                                                    pass
                                        
                                        # Verificación final: esperar a que aparezca el campo nenvio
                                        try:
                                            WebDriverWait(driver, 10).until(
                                                EC.presence_of_element_located((By.NAME, "nenvio"))
                                            )
                                            print("[OK] Campo 'nenvio' encontrado, formulario cargado correctamente")
                                        except:
                                            print("[WARNING] Campo 'nenvio' no encontrado aún, pero continuando...")
                                            
                                    except Exception as frame_error:
                                        print(f"[WARNING] Error al cambiar de frame: {frame_error}")
                                        import traceback
                                        traceback.print_exc()
                                        print("   Intentando continuar en el frame actual...")
                                    
                                except Exception as e:
                                    print(f"Error al hacer clic: {e}")
                                    try:
                                        driver.execute_script("arguments[0].click();", resultado_78)
                                        print("[OK] Clic realizado mediante JavaScript")
                                        time.sleep(1.5)  # Optimizado: reducido de 5s a 1.5s
                                        
                                        # Después del clic, buscar el frame correcto
                                        print("\nBuscando el frame correcto después del clic...")
                                        try:
                                            # Verificar si el formulario está en el frame actual
                                            WebDriverWait(driver, 5).until(
                                                EC.presence_of_element_located((By.NAME, "nenvio"))
                                            )
                                            print("[OK] Formulario encontrado en el frame actual")
                                        except:
                                            # Buscar en otros frames
                                            driver.switch_to.default_content()
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            for frame in frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    if frame_name:
                                                        driver.switch_to.frame(frame_name)
                                                        time.sleep(2)
                                                        try:
                                                            WebDriverWait(driver, 5).until(
                                                                EC.presence_of_element_located((By.NAME, "nenvio"))
                                                            )
                                                            print(f"[OK] Formulario encontrado en el frame '{frame_name}'")
                                                            break
                                                        except:
                                                            driver.switch_to.default_content()
                                                except:
                                                    continue
                                    except Exception as e2:
                                        print(f"Error al hacer clic con JavaScript: {e2}")
                            else:
                                print("[WARNING] No se encontró el resultado '7.8 Consulta a Histórico' en los resultados de búsqueda.")
                                print("   El navegador permanecerá abierto para que puedas revisar manualmente.")
                                return
                                
                        except Exception as e:
                            print(f"Error al usar el buscador: {e}")
                            import traceback
                            traceback.print_exc()
                            return
                    else:
                        print("[WARNING] No se pudo encontrar el campo de búsqueda.")
                        print("   El navegador permanecerá abierto para que puedas navegar manualmente.")
                        print("   No se puede continuar sin el campo de búsqueda.")
                        # No continuar con el resto del flujo si no se encuentra el buscador
                        return
                    
                    # ================================================================
                    # === DESCARGA MASIVA DE GUÍAS DESDE ARCHIVO EXCEL ===
                    # ================================================================
                    ruta_excel = os.path.join(WEBSITE_DIR, "GUIAS.xlsx")
                    ruta_excel_backup = os.path.join(WEBSITE_DIR, "GUIAS_RESULTADO.xlsx")
                    print("\nLeyendo archivo de guías desde Excel...")
                    wb_guias = openpyxl.load_workbook(ruta_excel)
                    ws_guias = wb_guias.active
                    lista_guias = []
                    for row_excel in range(2, ws_guias.max_row + 1):
                        val = ws_guias.cell(row=row_excel, column=1).value
                        if val and str(val).strip():
                            lista_guias.append((row_excel, str(val).strip()))

                    print(f"\n[OK] Se encontraron {len(lista_guias)} guías en el archivo Excel")
                    for g_row, g_num in lista_guias:
                        print(f"   - {g_num}")

                    guias_ok = 0
                    guias_error = 0
                    total_guias = len(lista_guias)
                    # La navegación de vuelta se hace re-clickeando 7.8 en el menú

                    for idx_guia, (fila_excel, guia_prueba) in enumerate(lista_guias):
                        print(f"\n{"=" * 60}")
                        print(f"PROCESANDO GUÍA {idx_guia + 1} de {total_guias}: {guia_prueba}")
                        print(f"{"=" * 60}")
                        guia_exitosa = False
                        try:
                            # Buscar el campo 'nenvio' usando un método completamente diferente
                            print("\n[PASO 3] Buscando el campo 'Envío' (nenvio) para escribir la guía...")
                    
                            # Esperar a que cargue el formulario (optimizado)
                            tiempo_espera = 0.5 if idx_guia > 0 else 1.5  # Reducido significativamente
                            print(f"   Esperando {tiempo_espera}s a que cargue el formulario...")
                            time.sleep(tiempo_espera)
                    
                            # MÉTODO NUEVO: Usar JavaScript para buscar agresivamente en TODOS los contextos
                            print("   Usando JavaScript para buscar el campo en todos los contextos...")
                            resultado_js = driver.execute_script("""
                                var guia = arguments[0];
                                var resultados = [];
                        
                                // Función recursiva para buscar en todos los frames
                                function buscarEnFrames(doc, nivel) {
                                    if (nivel > 10) return; // Evitar recursión infinita
                            
                                    try {
                                        // Buscar campo nenvio en este documento
                                        var campo = doc.querySelector('input[name="nenvio"]');
                                        if (campo && campo.offsetParent !== null) {
                                            return {exito: true, campo: campo, contexto: 'documento nivel ' + nivel};
                                        }
                                
                                        // Buscar en todos los frames/iframes
                                        var frames = doc.querySelectorAll('frame, iframe');
                                        for (var i = 0; i < frames.length; i++) {
                                            try {
                                                var frameDoc = frames[i].contentDocument || frames[i].contentWindow.document;
                                                var resultado = buscarEnFrames(frameDoc, nivel + 1);
                                                if (resultado && resultado.exito) {
                                                    return resultado;
                                                }
                                            } catch(e) {
                                                continue;
                                            }
                                        }
                                    } catch(e) {
                                        return null;
                                    }
                                    return null;
                                }
                        
                                // Buscar en el documento principal
                                var resultado = buscarEnFrames(document, 0);
                                if (resultado && resultado.exito) {
                                    var campo = resultado.campo;
                                    // Limpiar completamente el campo
                                    campo.focus();
                                    campo.select();
                                    campo.value = '';
                                    campo.dispatchEvent(new Event('input', { bubbles: true }));
                                    campo.dispatchEvent(new Event('change', { bubbles: true }));
                                    // Escribir el nuevo valor
                                    campo.value = guia;
                                    campo.dispatchEvent(new Event('input', { bubbles: true }));
                                    campo.dispatchEvent(new Event('change', { bubbles: true }));
                                    campo.dispatchEvent(new Event('blur', { bubbles: true }));
                                    return {exito: true, campo: 'nenvio', valor: campo.value, contexto: resultado.contexto};
                                }
                        
                                // Si no se encontró nenvio, listar TODOS los inputs para debug
                                function listarInputs(doc, nivel, prefijo) {
                                    var inputs = [];
                                    try {
                                        var todosInputs = doc.querySelectorAll('input[type="text"]');
                                        for (var i = 0; i < todosInputs.length; i++) {
                                            if (todosInputs[i].offsetParent !== null) {
                                                inputs.push({
                                                    name: todosInputs[i].name || '',
                                                    id: todosInputs[i].id || '',
                                                    maxlength: todosInputs[i].maxlength || '',
                                                    value: todosInputs[i].value || '',
                                                    contexto: prefijo + ' nivel ' + nivel
                                                });
                                            }
                                        }
                                
                                        // Buscar en frames
                                        var frames = doc.querySelectorAll('frame, iframe');
                                        for (var f = 0; f < frames.length; f++) {
                                            try {
                                                var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                var frameInputs = listarInputs(frameDoc, nivel + 1, prefijo + '->frame[' + f + ']');
                                                inputs = inputs.concat(frameInputs);
                                            } catch(e) {
                                                continue;
                                            }
                                        }
                                    } catch(e) {
                                        // Ignorar errores de acceso
                                    }
                                    return inputs;
                                }
                        
                                var todosInputs = listarInputs(document, 0, 'doc');
                                return {exito: false, inputs: todosInputs, mensaje: 'Campo nenvio no encontrado'};
                            """, guia_prueba)
                    
                            campo_envio = None
                    
                            if resultado_js and resultado_js.get('exito'):
                                print(f"[OK] Campo encontrado usando JavaScript: {resultado_js.get('contexto', 'desconocido')}")
                                print(f"[OK] Valor escrito: {resultado_js.get('valor', '')}")
                                campo_envio = True  # Marcar como encontrado
                                # Si JavaScript ya escribió la guía, saltar directamente al PASO 4 sin buscar con Selenium
                                print("[OK] Guía escrita exitosamente. Continuando al siguiente paso...")
                            else:
                                # Mostrar todos los inputs encontrados para debug
                                if resultado_js and 'inputs' in resultado_js:
                                    inputs_encontrados = resultado_js.get('inputs', [])
                                    print(f"   [WARNING] Campo 'nenvio' no encontrado. Inputs disponibles ({len(inputs_encontrados)}):")
                                    for idx, inp in enumerate(inputs_encontrados[:20]):  # Mostrar los primeros 20
                                        print(f"      Input #{idx+1}: name='{inp.get('name', '')}', id='{inp.get('id', '')}', maxlength='{inp.get('maxlength', '')}', contexto='{inp.get('contexto', '')}'")
                        
                                # Intentar buscar el campo usando Selenium después de ver los inputs
                                print("   Intentando buscar el campo usando Selenium...")
                        
                                # PRIMERO: Intentar encontrar el campo por name='nenvio' con espera
                                print("   Intentando encontrar campo por name='nenvio' con espera...")
                        
                                # Buscar en todas las ventanas disponibles
                                todas_ventanas = driver.window_handles
                                print(f"   Encontradas {len(todas_ventanas)} ventanas")
                        
                                for ventana_idx, ventana in enumerate(todas_ventanas):
                                    try:
                                        driver.switch_to.window(ventana)
                                        print(f"   Probando ventana #{ventana_idx+1}...")
                                
                                        # Intentar en el documento principal
                                        try:
                                            driver.switch_to.default_content()
                                            campo_envio = WebDriverWait(driver, 10).until(
                                                EC.presence_of_element_located((By.NAME, "nenvio"))
                                            )
                                            if campo_envio.is_displayed():
                                                print(f"[OK] Campo 'nenvio' encontrado en ventana #{ventana_idx+1} (documento principal)")
                                                break
                                        except:
                                            pass
                                
                                        # Si no se encontró, buscar en todos los frames de esta ventana
                                        if not campo_envio:
                                            driver.switch_to.default_content()
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                            todos_frames = list(frames) + list(iframes)
                                    
                                            print(f"      Encontrados {len(todos_frames)} frames/iframes en esta ventana")
                                    
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                            
                                                    driver.switch_to.default_content()
                                                    if frame_name:
                                                        driver.switch_to.frame(frame_name)
                                                    elif frame_id:
                                                        driver.switch_to.frame(frame_id)
                                                    else:
                                                        continue
                                            
                                                    try:
                                                        campo_envio = WebDriverWait(driver, 5).until(
                                                            EC.presence_of_element_located((By.NAME, "nenvio"))
                                                        )
                                                        if campo_envio.is_displayed():
                                                            print(f"[OK] Campo 'nenvio' encontrado en ventana #{ventana_idx+1}, frame '{frame_name or frame_id}'")
                                                            break
                                                    except:
                                                        continue
                                            
                                                    if campo_envio:
                                                        break
                                                except Exception as e:
                                                    print(f"      Error al cambiar al frame: {e}")
                                                    continue
                                
                                        if campo_envio:
                                            break
                                    except Exception as e:
                                        print(f"   Error al cambiar a ventana #{ventana_idx+1}: {e}")
                                        continue
                        
                                # Si no se encontró por name, usar selectores priorizados (similar a funcionalidad_codigo)
                                if not campo_envio:
                                    print("   Campo no encontrado por name, intentando otras estrategias...")
                            
                                    # Estrategias priorizadas: primero buscar el campo específico por name y atributos
                                    selectores_envio = [
                                        # Prioridad 1: Campo específico por name (más confiable)
                                        (By.NAME, "nenvio"),
                                        (By.XPATH, "//input[@name='nenvio']"),
                                        # Prioridad 2: Campo con maxlength='12' (característica del campo nenvio)
                                        (By.XPATH, "//input[@name='nenvio' and @maxlength='12']"),
                                        (By.XPATH, "//input[@type='text' and @maxlength='12']"),
                                        # Prioridad 3: Inputs de texto que NO sean funcionalidad_codigo
                                        (By.XPATH, "//input[@type='text' and @name!='funcionalidad_codigo' and @name!='']"),
                                        # Prioridad 4: Buscar por etiquetas cercanas que contengan "Envío" o "Envio"
                                        (By.XPATH, "//label[contains(translate(text(),'ENVIO','envio'),'envio')]/following-sibling::input[1]"),
                                        (By.XPATH, "//label[contains(translate(text(),'ENVIO','envio'),'envio')]/../input"),
                                        (By.XPATH, "//td[contains(translate(text(),'ENVIO','envio'),'envio')]/following::input[1]"),
                                        # Prioridad 5: Inputs de texto genéricos (último recurso)
                                        (By.XPATH, "//input[@type='text']"),
                                    ]
                            
                                    # Buscar en la ventana actual y todos sus frames
                                    for ventana_idx2, ventana2 in enumerate(todas_ventanas):
                                        try:
                                            driver.switch_to.window(ventana2)
                                            driver.switch_to.default_content()
                                    
                                            for by, selector in selectores_envio:
                                                try:
                                                    # Buscar sin imprimir mensajes de timeout
                                                    campo_envio = WebDriverWait(driver, 5).until(
                                                        EC.presence_of_element_located((by, selector))
                                                    )
                                                    if campo_envio.is_displayed():
                                                        # VERIFICACIÓN IMPORTANTE: Asegurarse de que NO sea el campo de búsqueda del menú
                                                        campo_name = campo_envio.get_attribute("name") or ""
                                                        campo_id = campo_envio.get_attribute("id") or ""
                                                
                                                        print(f"      Campo encontrado: name='{campo_name}', id='{campo_id}'")
                                                
                                                        # Si es el campo de búsqueda del menú (funcionalidad_codigo), ignorarlo
                                                        if campo_name == "funcionalidad_codigo" or "funcionalidad" in campo_name.lower():
                                                            print(f"      [WARNING] Campo es el buscador del menú, ignorando")
                                                            campo_envio = None
                                                            continue
                                                
                                                        # Verificar que sea el campo correcto (nenvio)
                                                        if campo_name == "nenvio" or "nenvio" in campo_name.lower():
                                                            print(f"[OK] Campo 'Envío' (nenvio) encontrado: {by} = '{selector}', name='{campo_name}'")
                                                            break
                                                        else:
                                                            # Si no es nenvio pero tiene maxlength=12, podría ser el correcto
                                                            maxlength = campo_envio.get_attribute("maxlength") or ""
                                                            if maxlength == "12":
                                                                print(f"[OK] Campo 'Envío' encontrado (por maxlength=12): {by} = '{selector}', name='{campo_name}'")
                                                                break
                                                            else:
                                                                print(f"      [WARNING] Campo no es nenvio: name='{campo_name}', maxlength='{maxlength}'")
                                                                campo_envio = None
                                                                continue
                                                except TimeoutException:
                                                    continue
                                                except NoSuchElementException:
                                                    continue
                                                except Exception as e:
                                                    print(f"      [WARNING] Error: {e}")
                                                    continue
                                    
                                            if campo_envio:
                                                break
                                    
                                            # Si no se encontró en el documento principal, buscar en frames
                                            driver.switch_to.default_content()
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                            todos_frames = list(frames) + list(iframes)
                                    
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                            
                                                    if frame_name or frame_id:
                                                        driver.switch_to.default_content()
                                                        if frame_name:
                                                            driver.switch_to.frame(frame_name)
                                                        else:
                                                            driver.switch_to.frame(frame_id)
                                                
                                                        for by, selector in selectores_envio:
                                                            try:
                                                                campo_envio = WebDriverWait(driver, 3).until(
                                                                    EC.presence_of_element_located((by, selector))
                                                                )
                                                                if campo_envio.is_displayed():
                                                                    campo_name = campo_envio.get_attribute("name") or ""
                                                                    if campo_name == "funcionalidad_codigo":
                                                                        campo_envio = None
                                                                        continue
                                                                    if campo_name == "nenvio" or (campo_envio.get_attribute("maxlength") == "12" and campo_name != "funcionalidad_codigo"):
                                                                        print(f"[OK] Campo 'Envío' encontrado en frame '{frame_name or frame_id}'")
                                                                        break
                                                            except:
                                                                continue
                                                
                                                        if campo_envio:
                                                            break
                                                except:
                                                    continue
                                    
                                            if campo_envio:
                                                break
                                        except:
                                            continue
                        
                                # Si encontramos el campo con Selenium, escribir la guía
                                if campo_envio and campo_envio is not True:
                                    try:
                                        print(f"[OK] Escribiendo la guía '{guia_prueba}' en el campo encontrado...")
                                
                                        # Hacer scroll al campo
                                        driver.execute_script("arguments[0].scrollIntoView(true);", campo_envio)
                                        time.sleep(0.3)  # Optimizado: reducido de 0.5 a 0.3
                                
                                        # Hacer foco en el campo
                                        campo_envio.click()
                                        time.sleep(0.2)  # Optimizado: reducido de 0.3 a 0.2
                                
                                        # Limpiar y escribir la guía
                                        campo_envio.clear()
                                        time.sleep(0.2)  # Optimizado: reducido de 0.3 a 0.2
                                        campo_envio.send_keys(guia_prueba)
                                        time.sleep(0.3)  # Optimizado: reducido de 0.5 a 0.3
                                
                                        # Verificar que se escribió correctamente
                                        valor_escrito = campo_envio.get_attribute("value") or campo_envio.text
                                        if valor_escrito == guia_prueba:
                                            print(f"[OK] Número de guía ingresado correctamente: {guia_prueba}")
                                        else:
                                            print(f"[WARNING] Valor escrito: '{valor_escrito}', esperado: '{guia_prueba}'")
                                            # Intentar de nuevo
                                            campo_envio.clear()
                                            campo_envio.send_keys(guia_prueba)
                                            print(f"[OK] Número de guía reingresado: {guia_prueba}")
                                        time.sleep(0.3)  # Optimizado: reducido de 0.5 a 0.3
                                    except Exception as e:
                                        print(f"Error al escribir en el campo: {e}")
                                        import traceback
                                        traceback.print_exc()
                                        campo_envio = None
                    
                            # Verificar si se encontró y escribió el campo
                            if not campo_envio:
                                print("[WARNING] No se pudo encontrar el campo 'Envío' (nenvio).")
                                print("   El navegador permanecerá abierto para que puedas revisar manualmente.")
                                print("   Revisa los inputs listados arriba para identificar el campo correcto.")
                                continue
                    
                            # Selectores priorizados: primero el campo específico por name
                            selectores_envio = [
                                # Prioridad 1: Campo específico por name (más confiable) - nenvio
                                (By.NAME, "nenvio"),
                                (By.XPATH, "//input[@name='nenvio']"),
                                (By.XPATH, "//input[@name='nenvio' and @maxlength='12']"),  # El campo tiene maxlength="12"
                                # Prioridad 2: Otros nombres comunes
                                (By.NAME, "envio"),
                                (By.ID, "envio"),
                                (By.NAME, "Envío"),
                                (By.ID, "Envío"),
                                (By.XPATH, "//input[contains(translate(@name,'ENVIO','envio'),'envio')]"),
                                (By.XPATH, "//input[contains(translate(@id,'ENVIO','envio'),'envio')]"),
                                (By.XPATH, "//input[contains(translate(@placeholder,'ENVIO','envio'),'envio')]"),
                                # Prioridad 3: Buscar por etiquetas cercanas
                                (By.XPATH, "//label[contains(translate(text(),'ENVIO','envio'),'envio')]/following-sibling::input"),
                                (By.XPATH, "//label[contains(translate(text(),'ENVIO','envio'),'envio')]/../input"),
                                (By.XPATH, "//td[contains(translate(text(),'ENVIO','envio'),'envio')]/following::input[1]"),
                                # Prioridad 4: Inputs de texto genéricos
                                (By.XPATH, "//input[@type='text' and @maxlength='12']"),  # El campo tiene maxlength="12"
                                (By.XPATH, "//input[@type='text']"),
                                (By.XPATH, "//input[not(@type='hidden') and not(@type='submit') and not(@type='button')]"),
                            ]

                            for by, selector in (selectores_envio if not campo_envio else []):
                                try:
                                    # Buscar sin imprimir mensajes de timeout
                                    campo_envio = WebDriverWait(driver, 5).until(  # Reducido de 10 a 5 segundos
                                        EC.presence_of_element_located((by, selector))
                                    )
                                    if campo_envio.is_displayed():
                                        # VERIFICACIÓN IMPORTANTE: Asegurarse de que NO sea el campo de búsqueda del menú
                                        campo_name = campo_envio.get_attribute("name") or ""
                                        campo_id = campo_envio.get_attribute("id") or ""
                                
                                        print(f"      Campo encontrado: name='{campo_name}', id='{campo_id}'")
                                
                                        # Si es el campo de búsqueda del menú (funcionalidad_codigo), ignorarlo
                                        if campo_name == "funcionalidad_codigo" or "funcionalidad" in campo_name.lower():
                                            print(f"      [WARNING] Campo es el buscador del menú, ignorando")
                                            campo_envio = None
                                            continue
                                
                                        # Verificar que sea el campo correcto (nenvio)
                                        if campo_name == "nenvio" or "nenvio" in campo_name.lower():
                                            print(f"[OK] Campo 'Envío' (nenvio) encontrado: {by} = '{selector}', name='{campo_name}'")
                                            break
                                        else:
                                            # Si no es nenvio pero tiene maxlength=12, podría ser el correcto
                                            maxlength = campo_envio.get_attribute("maxlength") or ""
                                            if maxlength == "12":
                                                print(f"[OK] Campo 'Envío' encontrado (por maxlength=12): {by} = '{selector}', name='{campo_name}'")
                                                break
                                            else:
                                                print(f"      [WARNING] Campo no es nenvio: name='{campo_name}', maxlength='{maxlength}'")
                                                campo_envio = None
                                                continue
                                    else:
                                        print(f"      [WARNING] Campo encontrado pero no está visible")
                                        campo_envio = None
                                except TimeoutException:
                                    # Timeout silencioso - continuar con el siguiente selector
                                    continue
                                except NoSuchElementException:
                                    print(f"      [WARNING] Elemento no encontrado")
                                    continue
                                except Exception as e:
                                    print(f"      [WARNING] Error: {e}")
                                    continue

                            if campo_envio is None:
                                print("[WARNING] No se pudo localizar el campo 'Envío' en el frame actual.")
                                print("   Intentando buscar en otros frames...")
                                try:
                                    # Volver al frame principal y buscar en todos los frames
                                    driver.switch_to.default_content()
                                    frames = driver.find_elements(By.TAG_NAME, "frame")
                                    iframes = driver.find_elements(By.TAG_NAME, "iframe")
                            
                                    for frame in frames:
                                        try:
                                            frame_name = frame.get_attribute("name") or ""
                                            if frame_name and frame_name != "menu":
                                                driver.switch_to.frame(frame_name)
                                                time.sleep(1)
                                                try:
                                                    campo_envio = WebDriverWait(driver, 5).until(
                                                        EC.presence_of_element_located((By.NAME, "nenvio"))
                                                    )
                                                    if campo_envio.is_displayed():
                                                        print(f"[OK] Campo 'Envío' encontrado en el frame '{frame_name}'")
                                                        break
                                                except:
                                                    driver.switch_to.default_content()
                                                    continue
                                        except:
                                            try:
                                                driver.switch_to.default_content()
                                            except:
                                                pass
                            
                                    # Si aún no se encuentra, intentar búsqueda por etiquetas
                                    if campo_envio is None:
                                        print("   Intentando búsqueda por etiquetas cercanas...")
                                        try:
                                            labels = driver.find_elements(By.XPATH, "//label | //td | //span | //div")
                                            for label in labels:
                                                texto_label = label.text.strip().upper()
                                                if "ENVIO" in texto_label or "ENVÍO" in texto_label:
                                                    try:
                                                        campo_envio = label.find_element(By.XPATH, ".//following::input[1] | .//../input | .//input")
                                                        print(f"[OK] Campo 'Envío' encontrado cerca del label: '{label.text}'")
                                                        break
                                                    except:
                                                        continue
                                        except Exception as e:
                                            print(f"Error en búsqueda por etiquetas: {e}")
                                except Exception as e:
                                    print(f"Error al buscar en otros frames: {e}")
                    
                            if campo_envio:
                                # Si campo_envio es True, significa que ya se escribió directamente en el elemento activo
                                if campo_envio is True:
                                    print("   La guía ya fue escrita directamente en el elemento activo")
                                    time.sleep(1)  # Esperar un momento
                                else:
                                    # Si campo_envio es un WebElement, escribir normalmente
                                    try:
                                        # Hacer scroll al campo
                                        driver.execute_script("arguments[0].scrollIntoView(true);", campo_envio)
                                        time.sleep(0.5)
                                
                                        # Limpiar el campo y escribir la guía
                                        campo_envio.click()  # Hacer foco en el campo
                                        time.sleep(0.3)
                                        campo_envio.clear()
                                        time.sleep(0.3)
                                        campo_envio.send_keys(guia_prueba)
                                        time.sleep(0.5)
                                
                                        # Verificar que se escribió correctamente
                                        valor_escrito = campo_envio.get_attribute("value") or campo_envio.text
                                        if valor_escrito == guia_prueba:
                                            print(f"[OK] Número de guía ingresado correctamente: {guia_prueba}")
                                        else:
                                            print(f"[WARNING] Valor escrito: '{valor_escrito}', esperado: '{guia_prueba}'")
                                            # Intentar de nuevo
                                            campo_envio.clear()
                                            campo_envio.send_keys(guia_prueba)
                                            print(f"[OK] Número de guía reingresado: {guia_prueba}")
                                        time.sleep(0.5)
                                    except Exception as e:
                                        print(f"Error al escribir en el campo: {e}")
                    
                            # Buscar el botón "BUSCAR" usando JavaScript y hacer clic directamente
                            # Si llegamos aquí, la guía fue escrita exitosamente
                            print("\n[PASO 4] Buscando y haciendo clic en el botón 'BUSCAR' usando JavaScript...")
                            time.sleep(2)  # Esperar un momento después de escribir la guía
                    
                            boton_buscar = None
                    
                            # MÉTODO NUEVO: Usar JavaScript para buscar el botón por id="find" y hacer clic directamente
                            print("   Buscando el botón por id='find' usando JavaScript en todos los contextos...")
                            resultado_click = driver.execute_script("""
                                // Función recursiva para buscar el botón en todos los frames
                                function buscarYBotonClick(doc, nivel) {
                                    if (nivel > 10) return false; // Evitar recursión infinita
                            
                                    try {
                                        // PRIMERO: Buscar por id="find" (más específico)
                                        var boton = doc.getElementById('find');
                                        if (boton && boton.offsetParent !== null) {
                                            // Hacer clic directamente
                                            boton.click();
                                            return {exito: true, metodo: 'id=find', nivel: nivel};
                                        }
                                
                                        // SEGUNDO: Buscar por class="btn3"
                                        var botonesBtn3 = doc.querySelectorAll('button.btn3, button[class*="btn3"]');
                                        for (var i = 0; i < botonesBtn3.length; i++) {
                                            if (botonesBtn3[i].offsetParent !== null) {
                                                botonesBtn3[i].click();
                                                return {exito: true, metodo: 'class=btn3', nivel: nivel};
                                            }
                                        }
                                
                                        // TERCERO: Buscar botones con texto "Buscar"
                                        var botones = doc.querySelectorAll('button');
                                        for (var i = 0; i < botones.length; i++) {
                                            if (botones[i].offsetParent !== null) {
                                                var texto = (botones[i].textContent || '').trim().toUpperCase();
                                                if (texto.indexOf('BUSCAR') !== -1) {
                                                    botones[i].click();
                                                    return {exito: true, metodo: 'texto=Buscar', nivel: nivel};
                                                }
                                            }
                                        }
                                
                                        // Buscar en todos los frames/iframes
                                        var frames = doc.querySelectorAll('frame, iframe');
                                        for (var f = 0; f < frames.length; f++) {
                                            try {
                                                var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                var resultado = buscarYBotonClick(frameDoc, nivel + 1);
                                                if (resultado && resultado.exito) {
                                                    return resultado;
                                                }
                                            } catch(e) {
                                                continue;
                                            }
                                        }
                                    } catch(e) {
                                        return {exito: false, error: e.toString()};
                                    }
                                    return {exito: false};
                                }
                        
                                // Buscar en el documento principal
                                return buscarYBotonClick(document, 0);
                            """)
                    
                            if resultado_click and resultado_click.get('exito'):
                                metodo_usado = resultado_click.get('metodo', 'desconocido')
                                print(f"[OK] Botón 'BUSCAR' encontrado y clickeado usando: {metodo_usado}")
                                time.sleep(2)  # Optimizado: reducido de 5s a 2s
                                boton_buscar = True
                            else:
                                print("   [WARNING] No se encontró el botón con JavaScript, intentando con Selenium...")
                        
                                # Intentar con Selenium como respaldo
                                todas_ventanas = driver.window_handles
                                for ventana_idx, ventana in enumerate(todas_ventanas):
                                    try:
                                        driver.switch_to.window(ventana)
                                        driver.switch_to.default_content()
                                
                                        # Buscar por id="find"
                                        try:
                                            boton_buscar = WebDriverWait(driver, 3).until(
                                                EC.presence_of_element_located((By.ID, "find"))
                                            )
                                            if boton_buscar.is_displayed():
                                                print(f"[OK] Botón 'BUSCAR' encontrado por id='find' en ventana #{ventana_idx+1}")
                                                driver.execute_script("arguments[0].click();", boton_buscar)
                                                print("[OK] Clic realizado mediante JavaScript")
                                                time.sleep(3)
                                                break
                                        except:
                                            pass
                                
                                        # Buscar en frames
                                        if not boton_buscar:
                                            driver.switch_to.default_content()
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                            todos_frames = list(frames) + list(iframes)
                                    
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                            
                                                    if frame_name or frame_id:
                                                        driver.switch_to.default_content()
                                                        if frame_name:
                                                            driver.switch_to.frame(frame_name)
                                                        else:
                                                            driver.switch_to.frame(frame_id)
                                                
                                                        try:
                                                            boton_buscar = WebDriverWait(driver, 3).until(
                                                                EC.presence_of_element_located((By.ID, "find"))
                                                            )
                                                            if boton_buscar.is_displayed():
                                                                print(f"[OK] Botón 'BUSCAR' encontrado por id='find' en frame '{frame_name or frame_id}'")
                                                                driver.execute_script("arguments[0].click();", boton_buscar)
                                                                print("[OK] Clic realizado mediante JavaScript")
                                                                time.sleep(3)
                                                                break
                                                        except:
                                                            continue
                                        
                                                    if boton_buscar:
                                                        break
                                                except:
                                                    continue
                                
                                        if boton_buscar:
                                            break
                                    except:
                                        continue
                    
                            if not boton_buscar:
                                print("[WARNING] No se pudo encontrar ni hacer clic en el botón 'BUSCAR'.")
                                print("   El navegador permanecerá abierto para que puedas revisar manualmente.")
                                continue
                    
                            # PRIMERO: Buscar el botón usando los selectores específicos del HTML proporcionado
                            print("   Buscando el botón 'BUSCAR' usando selectores específicos...")
                    
                            # Selectores específicos basados en el HTML: <button class="btn3" id="find" onclick="javascript: seleccionar_busqueda(event);">Buscar</button>
                            selectores_especificos = [
                                (By.ID, "find"),  # Prioridad 1: Buscar por id="find"
                                (By.XPATH, "//button[@id='find']"),
                                (By.XPATH, "//button[@class='btn3']"),  # Prioridad 2: Buscar por class="btn3"
                                (By.XPATH, "//button[contains(@class,'btn3')]"),
                                (By.XPATH, "//button[contains(@onclick,'seleccionar_busqueda')]"),  # Prioridad 3: Buscar por onclick
                                (By.XPATH, "//button[contains(translate(text(),'BUSCAR','buscar'),'buscar')]"),  # Prioridad 4: Buscar por texto
                                (By.XPATH, "//button[normalize-space(text())='Buscar']"),
                            ]
                    
                            # Buscar en todas las ventanas y frames
                            todas_ventanas = driver.window_handles
                            for ventana_idx, ventana in enumerate(todas_ventanas):
                                try:
                                    driver.switch_to.window(ventana)
                                    driver.switch_to.default_content()
                            
                                    # Intentar con selectores específicos en el documento principal
                                    for by, selector in selectores_especificos:
                                        try:
                                            boton_buscar = WebDriverWait(driver, 3).until(
                                                EC.presence_of_element_located((by, selector))
                                            )
                                            if boton_buscar.is_displayed():
                                                texto_boton = (boton_buscar.text or "").strip()
                                                print(f"[OK] Botón 'BUSCAR' encontrado: {by} = '{selector}', texto='{texto_boton}'")
                                                break
                                        except:
                                            continue
                            
                                    if boton_buscar:
                                        break
                            
                                    # Si no se encontró, buscar en todos los frames
                                    driver.switch_to.default_content()
                                    frames = driver.find_elements(By.TAG_NAME, "frame")
                                    iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                    todos_frames = list(frames) + list(iframes)
                            
                                    for frame in todos_frames:
                                        try:
                                            frame_name = frame.get_attribute("name") or ""
                                            frame_id = frame.get_attribute("id") or ""
                                    
                                            if frame_name or frame_id:
                                                driver.switch_to.default_content()
                                                if frame_name:
                                                    driver.switch_to.frame(frame_name)
                                                else:
                                                    driver.switch_to.frame(frame_id)
                                        
                                                # Intentar con selectores específicos en este frame
                                                for by, selector in selectores_especificos:
                                                    try:
                                                        boton_buscar = WebDriverWait(driver, 3).until(
                                                            EC.presence_of_element_located((by, selector))
                                                        )
                                                        if boton_buscar.is_displayed():
                                                            texto_boton = (boton_buscar.text or "").strip()
                                                            print(f"[OK] Botón 'BUSCAR' encontrado en frame '{frame_name or frame_id}': {by} = '{selector}', texto='{texto_boton}'")
                                                            break
                                                    except:
                                                        continue
                                        
                                                if boton_buscar:
                                                    break
                                        except:
                                            continue
                            
                                    if boton_buscar:
                                        break
                                except:
                                    continue
                    
                            # Si no se encontró con selectores específicos, buscar en el mismo contexto donde se encontró el campo
                            if not boton_buscar and campo_envio and campo_envio is not True:
                                try:
                                    print("   Buscando el botón en el mismo contexto donde se encontró el campo...")
                                    # El campo_envio está en el contexto correcto, buscar botones cerca de él
                                    botones_cercanos = driver.find_elements(By.XPATH, "//button | //input[@type='submit'] | //input[@type='button']")
                                    print(f"   Encontrados {len(botones_cercanos)} botones en el contexto actual")
                            
                                    for boton in botones_cercanos:
                                        try:
                                            if boton.is_displayed():
                                                texto_boton = (boton.text or boton.get_attribute("value") or "").strip().upper()
                                                boton_id = boton.get_attribute("id") or ""
                                                boton_class = boton.get_attribute("class") or ""
                                                if "BUSCAR" in texto_boton or boton_id == "find" or "btn3" in boton_class:
                                                    boton_buscar = boton
                                                    print(f"[OK] Botón 'BUSCAR' encontrado en el mismo contexto: texto='{texto_boton}', id='{boton_id}'")
                                                    break
                                        except:
                                            continue
                                except Exception as e:
                                    print(f"   Error al buscar en el contexto actual: {e}")
                    
                                # Si no se encontró, usar JavaScript para buscar agresivamente en TODOS los contextos
                                print("   Usando JavaScript para buscar el botón en todos los contextos...")
                                resultado_boton_js = driver.execute_script("""
                                    var resultados = [];
                            
                                    // Función recursiva para buscar en todos los frames
                                    function buscarBotonEnFrames(doc, nivel) {
                                        if (nivel > 10) return null; // Evitar recursión infinita
                                
                                        try {
                                            // PRIMERO: Buscar por id="find" (más específico)
                                            var botonFind = doc.getElementById('find');
                                            if (botonFind && botonFind.offsetParent !== null) {
                                                return {
                                                    exito: true,
                                                    boton: botonFind,
                                                    texto: (botonFind.textContent || '').trim(),
                                                    id: botonFind.id || '',
                                                    name: botonFind.name || '',
                                                    contexto: 'documento nivel ' + nivel + ' (por id=find)'
                                                };
                                            }
                                    
                                            // SEGUNDO: Buscar por class="btn3"
                                            var botonesBtn3 = doc.querySelectorAll('button.btn3, button[class*="btn3"]');
                                            for (var b = 0; b < botonesBtn3.length; b++) {
                                                if (botonesBtn3[b].offsetParent !== null) {
                                                    return {
                                                        exito: true,
                                                        boton: botonesBtn3[b],
                                                        texto: (botonesBtn3[b].textContent || '').trim(),
                                                        id: botonesBtn3[b].id || '',
                                                        name: botonesBtn3[b].name || '',
                                                        contexto: 'documento nivel ' + nivel + ' (por class=btn3)'
                                                    };
                                                }
                                            }
                                    
                                            // TERCERO: Buscar botones que contengan "BUSCAR" o "buscar" en el texto
                                            var botones = doc.querySelectorAll('button, input[type="submit"], input[type="button"]');
                                            for (var i = 0; i < botones.length; i++) {
                                                var boton = botones[i];
                                                if (boton.offsetParent !== null) { // Solo botones visibles
                                                    var texto = (boton.textContent || boton.value || '').trim().toUpperCase();
                                                    var id = boton.id || '';
                                                    var name = boton.name || '';
                                                    var onclick = boton.getAttribute('onclick') || '';
                                            
                                                    // Verificar si el texto contiene "BUSCAR" o tiene onclick con "seleccionar_busqueda"
                                                    if (texto.indexOf('BUSCAR') !== -1 || 
                                                        id === 'find' ||
                                                        id.toUpperCase().indexOf('BUSCAR') !== -1 || 
                                                        name.toUpperCase().indexOf('BUSCAR') !== -1 ||
                                                        onclick.indexOf('seleccionar_busqueda') !== -1) {
                                                        return {
                                                            exito: true,
                                                            boton: boton,
                                                            texto: (boton.textContent || boton.value || '').trim(),
                                                            id: id,
                                                            name: name,
                                                            contexto: 'documento nivel ' + nivel
                                                        };
                                                    }
                                                }
                                            }
                                    
                                            // Buscar en todos los frames/iframes
                                            var frames = doc.querySelectorAll('frame, iframe');
                                            for (var f = 0; f < frames.length; f++) {
                                                try {
                                                    var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                    var resultado = buscarBotonEnFrames(frameDoc, nivel + 1);
                                                    if (resultado && resultado.exito) {
                                                        return resultado;
                                                    }
                                                } catch(e) {
                                                    continue;
                                                }
                                            }
                                        } catch(e) {
                                            return null;
                                        }
                                        return null;
                                    }
                            
                                    // Buscar en el documento principal
                                    var resultado = buscarBotonEnFrames(document, 0);
                                    if (resultado && resultado.exito) {
                                        return resultado;
                                    }
                            
                                    // Si no se encontró, listar TODOS los botones para debug
                                    function listarBotones(doc, nivel, prefijo) {
                                        var botones = [];
                                        try {
                                            var todosBotones = doc.querySelectorAll('button, input[type="submit"], input[type="button"]');
                                            for (var i = 0; i < todosBotones.length; i++) {
                                                if (todosBotones[i].offsetParent !== null) {
                                                    botones.push({
                                                        texto: (todosBotones[i].textContent || todosBotones[i].value || '').trim(),
                                                        id: todosBotones[i].id || '',
                                                        name: todosBotones[i].name || '',
                                                        type: todosBotones[i].type || '',
                                                        contexto: prefijo + ' nivel ' + nivel
                                                    });
                                                }
                                            }
                                    
                                            // Buscar en frames
                                            var frames = doc.querySelectorAll('frame, iframe');
                                            for (var f = 0; f < frames.length; f++) {
                                                try {
                                                    var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                    var frameBotones = listarBotones(frameDoc, nivel + 1, prefijo + '->frame[' + f + ']');
                                                    botones = botones.concat(frameBotones);
                                                } catch(e) {
                                                    continue;
                                                }
                                            }
                                        } catch(e) {
                                            // Ignorar errores de acceso
                                        }
                                        return botones;
                                    }
                            
                                    var todosBotones = listarBotones(document, 0, 'doc');
                                    return {exito: false, botones: todosBotones, mensaje: 'Botón BUSCAR no encontrado'};
                                """)
                        
                                if resultado_boton_js and resultado_boton_js.get('exito'):
                                    # JavaScript encontró el botón, hacer clic directamente
                                    print(f"[OK] Botón encontrado usando JavaScript: texto='{resultado_boton_js.get('texto', '')}', id='{resultado_boton_js.get('id', '')}'")
                                    try:
                                        # Hacer clic usando JavaScript
                                        driver.execute_script("""
                                            var boton = arguments[0];
                                            boton.click();
                                        """, resultado_boton_js.get('boton'))
                                        print("[OK] Clic realizado en el botón 'BUSCAR' mediante JavaScript")
                                        time.sleep(2)  # Optimizado: reducido de 3 a 2 segundos
                                        boton_buscar = True  # Marcar como encontrado
                                    except Exception as e:
                                        print(f"   Error al hacer clic con JavaScript: {e}")
                                        boton_buscar = None
                                else:
                                    # Mostrar todos los botones encontrados para debug
                                    if resultado_boton_js and 'botones' in resultado_boton_js:
                                        botones_encontrados = resultado_boton_js.get('botones', [])
                                        print(f"   [WARNING] Botón 'BUSCAR' no encontrado. Botones disponibles ({len(botones_encontrados)}):")
                                        for idx, btn in enumerate(botones_encontrados[:20]):  # Mostrar los primeros 20
                                            print(f"      Botón #{idx+1}: texto='{btn.get('texto', '')}', id='{btn.get('id', '')}', name='{btn.get('name', '')}', type='{btn.get('type', '')}', contexto='{btn.get('contexto', '')}'")
                            
                                    # Intentar buscar usando Selenium en todas las ventanas y frames
                                    print("   Intentando buscar el botón usando Selenium...")
                                    todas_ventanas = driver.window_handles
                            
                                    for ventana_idx, ventana in enumerate(todas_ventanas):
                                        try:
                                            driver.switch_to.window(ventana)
                                            driver.switch_to.default_content()
                                    
                                            # Buscar todos los botones visibles
                                            try:
                                                todos_botones = driver.find_elements(By.XPATH, "//button | //input[@type='submit'] | //input[@type='button']")
                                                print(f"   Encontrados {len(todos_botones)} botones en ventana #{ventana_idx+1}")
                                        
                                                for boton in todos_botones:
                                                    try:
                                                        if boton.is_displayed():
                                                            texto_boton = (boton.text or boton.get_attribute("value") or "").strip().upper()
                                                            boton_id = boton.get_attribute("id") or ""
                                                            boton_name = boton.get_attribute("name") or ""
                                                    
                                                            # Verificar si el texto contiene "BUSCAR"
                                                            if "BUSCAR" in texto_boton:
                                                                boton_buscar = boton
                                                                print(f"[OK] Botón 'BUSCAR' encontrado: texto='{texto_boton}', id='{boton_id}', name='{boton_name}'")
                                                                break
                                                    except:
                                                        continue
                                        
                                                if boton_buscar:
                                                    break
                                            except:
                                                pass
                                    
                                            # Buscar en frames
                                            if not boton_buscar:
                                                driver.switch_to.default_content()
                                                frames = driver.find_elements(By.TAG_NAME, "frame")
                                                iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                                todos_frames = list(frames) + list(iframes)
                                        
                                                for frame in todos_frames:
                                                    try:
                                                        frame_name = frame.get_attribute("name") or ""
                                                        frame_id = frame.get_attribute("id") or ""
                                                
                                                        if frame_name or frame_id:
                                                            driver.switch_to.default_content()
                                                            if frame_name:
                                                                driver.switch_to.frame(frame_name)
                                                            else:
                                                                driver.switch_to.frame(frame_id)
                                                    
                                                            try:
                                                                todos_botones = driver.find_elements(By.XPATH, "//button | //input[@type='submit'] | //input[@type='button']")
                                                                for boton in todos_botones:
                                                                    try:
                                                                        if boton.is_displayed():
                                                                            texto_boton = (boton.text or boton.get_attribute("value") or "").strip().upper()
                                                                            if "BUSCAR" in texto_boton:
                                                                                boton_buscar = boton
                                                                                print(f"[OK] Botón 'BUSCAR' encontrado en frame '{frame_name or frame_id}'")
                                                                                break
                                                                    except:
                                                                        continue
                                                        
                                                                if boton_buscar:
                                                                    break
                                                            except:
                                                                continue
                                                    except:
                                                        continue
                                    
                                            if boton_buscar:
                                                break
                                        except:
                                            continue
                    
                            if boton_buscar:
                                # Si boton_buscar es True, JavaScript ya hizo clic
                                if boton_buscar is True:
                                    print("[OK] El botón 'BUSCAR' ya fue clickeado por JavaScript. Continuando...")
                                else:
                                    # Si boton_buscar es un WebElement, hacer clic con Selenium
                                    try:
                                        # Hacer scroll al botón
                                        driver.execute_script("arguments[0].scrollIntoView(true);", boton_buscar)
                                        time.sleep(0.5)
                                
                                        # Intentar hacer clic
                                        boton_buscar.click()
                                        print("[OK] Clic realizado en el botón 'Buscar'")
                                        time.sleep(3)
                                    except Exception as e:
                                        print(f"Error al hacer clic en el botón: {e}")
                                        try:
                                            # Intentar clic con JavaScript
                                            driver.execute_script("arguments[0].click();", boton_buscar)
                                            print("[OK] Clic realizado mediante JavaScript")
                                            time.sleep(3)
                                        except Exception as e2:
                                            print(f"Error al hacer clic con JavaScript: {e2}")
                            else:
                                print("[WARNING] No se encontró el botón 'Buscar', intentando con ENTER como respaldo...")
                                try:
                                    # Intentar enviar ENTER al elemento activo
                                    elemento_activo = driver.execute_script("return document.activeElement;")
                                    if elemento_activo:
                                        driver.execute_script("""
                                            var elem = arguments[0];
                                            var event = new KeyboardEvent('keydown', {
                                                key: 'Enter',
                                                code: 'Enter',
                                                keyCode: 13,
                                                which: 13,
                                                bubbles: true
                                            });
                                            elem.dispatchEvent(event);
                                        """, elemento_activo)
                                        print("[OK] ENTER enviado como respaldo")
                                        time.sleep(3)
                                except:
                                    print("[WARNING] No se pudo enviar ENTER")
                                    time.sleep(3)
                    
                            # PASO 5: Buscar el enlace 'VA' usando JavaScript (basado en el HTML: <a class="btn2" href="javascript:goPOD('121483818');" title="Ver guía Entrega">VA</a>)
                            print("\n[PASO 5] Buscando enlace 'VA' (Ver guía de entrega) en los resultados...")
                            time.sleep(2)  # Esperar a que carguen los resultados después de buscar (optimizado)
                    
                            va_element = None
                    
                            # Usar JavaScript para buscar y hacer clic en el enlace VA directamente (evita stale element)
                            print("   Buscando y haciendo clic en el enlace 'VA' usando JavaScript...")
                            resultado_va = driver.execute_script("""
                                // Función recursiva para buscar y hacer clic en el enlace VA
                                function buscarYClickVA(doc, nivel) {
                                    if (nivel > 10) return false; // Evitar recursión infinita
                            
                                    try {
                                        // PRIMERO: Buscar por class="btn2" y texto "VA"
                                        var enlaces = doc.querySelectorAll('a.btn2, a[class*="btn2"]');
                                        for (var i = 0; i < enlaces.length; i++) {
                                            if (enlaces[i].offsetParent !== null) {
                                                var texto = (enlaces[i].textContent || '').trim();
                                                if (texto === 'VA' || texto.toUpperCase() === 'VA') {
                                                    // Hacer clic directamente
                                                    enlaces[i].click();
                                                    return {exito: true, metodo: 'class=btn2', nivel: nivel};
                                                }
                                            }
                                        }
                                
                                        // SEGUNDO: Buscar cualquier enlace con texto "VA"
                                        var todosEnlaces = doc.querySelectorAll('a');
                                        for (var i = 0; i < todosEnlaces.length; i++) {
                                            if (todosEnlaces[i].offsetParent !== null) {
                                                var texto = (todosEnlaces[i].textContent || '').trim();
                                                if (texto === 'VA' || texto.toUpperCase() === 'VA') {
                                                    // Hacer clic directamente
                                                    todosEnlaces[i].click();
                                                    return {exito: true, metodo: 'texto=VA', nivel: nivel};
                                                }
                                            }
                                        }
                                
                                        // Buscar en todos los frames/iframes
                                        var frames = doc.querySelectorAll('frame, iframe');
                                        for (var f = 0; f < frames.length; f++) {
                                            try {
                                                var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                var resultado = buscarYClickVA(frameDoc, nivel + 1);
                                                if (resultado && resultado.exito) {
                                                    return resultado;
                                                }
                                            } catch(e) {
                                                continue;
                                            }
                                        }
                                    } catch(e) {
                                        return {exito: false, error: e.toString()};
                                    }
                                    return {exito: false};
                                }
                        
                                // Buscar y hacer clic en el documento principal
                                return buscarYClickVA(document, 0);
                            """)
                    
                            if resultado_va and resultado_va.get('exito'):
                                metodo_usado = resultado_va.get('metodo', 'desconocido')
                                print(f"[OK] Enlace 'VA' encontrado y clickeado usando: {metodo_usado}")
                                time.sleep(2)
                                va_element = True  # Marcar como encontrado
                            else:
                                # Intentar con Selenium como respaldo
                                print("   No se encontró con JavaScript, intentando con Selenium...")
                                selectores_va = [
                                    (By.XPATH, "//a[@class='btn2' and normalize-space(text())='VA']"),  # Prioridad 1: class="btn2" y texto "VA"
                                    (By.XPATH, "//a[contains(@class,'btn2') and normalize-space(text())='VA']"),
                                    (By.XPATH, "//a[normalize-space(text())='VA']"),  # Prioridad 2: cualquier enlace con texto "VA"
                                    (By.XPATH, "//a[contains(@title,'Ver guía Entrega')]"),  # Prioridad 3: por title
                                    (By.LINK_TEXT, "VA"),
                                ]
                        
                                todas_ventanas = driver.window_handles
                                for ventana_idx, ventana in enumerate(todas_ventanas):
                                    try:
                                        driver.switch_to.window(ventana)
                                        driver.switch_to.default_content()
                                
                                        for by, selector in selectores_va:
                                            try:
                                                va_element = WebDriverWait(driver, 5).until(
                                                    EC.presence_of_element_located((by, selector))
                                                )
                                                if va_element.is_displayed():
                                                    print(f"[OK] Enlace 'VA' encontrado: {by} = '{selector}'")
                                                    break
                                            except:
                                                continue
                                
                                        if va_element:
                                            break
                                
                                        # Buscar en frames
                                        driver.switch_to.default_content()
                                        frames = driver.find_elements(By.TAG_NAME, "frame")
                                        iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                        todos_frames = list(frames) + list(iframes)
                                
                                        for frame in todos_frames:
                                            try:
                                                frame_name = frame.get_attribute("name") or ""
                                                frame_id = frame.get_attribute("id") or ""
                                        
                                                if frame_name or frame_id:
                                                    driver.switch_to.default_content()
                                                    if frame_name:
                                                        driver.switch_to.frame(frame_name)
                                                    else:
                                                        driver.switch_to.frame(frame_id)
                                            
                                                    for by, selector in selectores_va:
                                                        try:
                                                            va_element = WebDriverWait(driver, 5).until(
                                                                EC.presence_of_element_located((by, selector))
                                                            )
                                                            if va_element.is_displayed():
                                                                print(f"[OK] Enlace 'VA' encontrado en frame '{frame_name or frame_id}': {by} = '{selector}'")
                                                                break
                                                        except:
                                                            continue
                                            
                                                    if va_element:
                                                        break
                                            except:
                                                continue
                                
                                        if va_element:
                                            break
                                    except:
                                        continue
                        
                                if va_element:
                                    # Si va_element es True, JavaScript ya hizo clic
                                    if va_element is True:
                                        print("[OK] El enlace 'VA' ya fue clickeado por JavaScript.")
                                    else:
                                        # Si va_element es un WebElement, hacer clic con Selenium
                                        try:
                                            print("[OK] Haciendo clic en 'VA' para abrir la POD...")
                                            driver.execute_script("arguments[0].click();", va_element)
                                            print("[OK] Clic en 'VA' realizado mediante JavaScript. Verifica la apertura/descarga de la POD.")
                                            time.sleep(2)
                                        except Exception as e:
                                            print(f"Error al hacer clic en VA: {e}")
                                            try:
                                                va_element.click()
                                                print("[OK] Clic en 'VA' realizado.")
                                                time.sleep(2)
                                            except Exception as e2:
                                                print(f"Error al hacer clic: {e2}")
                    
                            # Verificar si se encontró y clickeó el enlace VA
                            va_encontrado = (va_element and va_element is True) or (resultado_va and resultado_va.get('exito'))
                            if not va_encontrado:
                                print("[WARNING] No se encontró la sigla 'VA' para la guía indicada.")
                                print("   Es posible que esa guía no tenga POD asociada o que el formato de la tabla sea distinto.")
                            else:
                                # PASO 6: Buscar y hacer clic en el menú "P.O.D." (basado en: <a onclick="javascript:cambiar_contenido('t11', 't11base');">P.O.D.</a>)
                                print("\n[PASO 6] Buscando menú 'P.O.D.'...")
                                time.sleep(2)  # Esperar a que cargue la nueva página después del clic en VA
                        
                                pod_menu = None
                        
                                # Usar JavaScript para buscar y hacer clic en el enlace P.O.D. directamente
                                print("   Buscando y haciendo clic en el menú 'P.O.D.' usando JavaScript...")
                                resultado_pod = driver.execute_script("""
                                    // Función recursiva para buscar y hacer clic en el enlace P.O.D.
                                    function buscarYClickPOD(doc, nivel) {
                                        if (nivel > 10) return false; // Evitar recursión infinita
                                
                                        try {
                                            // Buscar enlaces con texto "P.O.D." o "POD"
                                            var enlaces = doc.querySelectorAll('a');
                                            for (var i = 0; i < enlaces.length; i++) {
                                                if (enlaces[i].offsetParent !== null) {
                                                    var texto = (enlaces[i].textContent || '').trim();
                                                    var onclick = enlaces[i].getAttribute('onclick') || '';
                                            
                                                    // Verificar si el texto contiene "P.O.D." o "POD" y si tiene el onclick con cambiar_contenido
                                                    if ((texto === 'P.O.D.' || texto.toUpperCase() === 'P.O.D.' || texto.toUpperCase() === 'POD') &&
                                                        onclick.indexOf('cambiar_contenido') !== -1) {
                                                        // Hacer clic directamente
                                                        enlaces[i].click();
                                                        return {exito: true, metodo: 'texto=P.O.D.', nivel: nivel};
                                                    }
                                                }
                                            }
                                    
                                            // Buscar en todos los frames/iframes
                                            var frames = doc.querySelectorAll('frame, iframe');
                                            for (var f = 0; f < frames.length; f++) {
                                                try {
                                                    var frameDoc = frames[f].contentDocument || frames[f].contentWindow.document;
                                                    var resultado = buscarYClickPOD(frameDoc, nivel + 1);
                                                    if (resultado && resultado.exito) {
                                                        return resultado;
                                                    }
                                                } catch(e) {
                                                    continue;
                                                }
                                            }
                                        } catch(e) {
                                            return {exito: false, error: e.toString()};
                                        }
                                        return {exito: false};
                                    }
                            
                                    // Buscar y hacer clic en el documento principal
                                    return buscarYClickPOD(document, 0);
                                """)
                        
                                if resultado_pod and resultado_pod.get('exito'):
                                    metodo_usado = resultado_pod.get('metodo', 'desconocido')
                                    print(f"[OK] Menú 'P.O.D.' encontrado y clickeado usando: {metodo_usado}")
                                    time.sleep(1)
                                    pod_menu = True
                                else:
                                    # Intentar con Selenium como respaldo
                                    print("   No se encontró con JavaScript, intentando con Selenium...")
                                    todas_ventanas = driver.window_handles
                            
                                    for ventana_idx, ventana in enumerate(todas_ventanas):
                                        try:
                                            driver.switch_to.window(ventana)
                                            driver.switch_to.default_content()
                                    
                                            # Buscar por texto "P.O.D." y onclick
                                            try:
                                                pod_menu = WebDriverWait(driver, 3).until(
                                                    EC.presence_of_element_located((By.XPATH, "//a[contains(text(),'P.O.D.') or contains(text(),'POD')]"))
                                                )
                                                if pod_menu.is_displayed():
                                                    onclick_attr = pod_menu.get_attribute('onclick') or ''
                                                    if 'cambiar_contenido' in onclick_attr:
                                                        print(f"[OK] Menú 'P.O.D.' encontrado en ventana #{ventana_idx+1}")
                                                        driver.execute_script("arguments[0].click();", pod_menu)
                                                        print("[OK] Clic realizado mediante JavaScript")
                                                        time.sleep(1)
                                                        break
                                                    else:
                                                        pod_menu = None
                                            except:
                                                pass
                                    
                                            if pod_menu:
                                                break
                                    
                                            # Buscar en frames
                                            driver.switch_to.default_content()
                                            frames = driver.find_elements(By.TAG_NAME, "frame")
                                            iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                            todos_frames = list(frames) + list(iframes)
                                    
                                            for frame in todos_frames:
                                                try:
                                                    frame_name = frame.get_attribute("name") or ""
                                                    frame_id = frame.get_attribute("id") or ""
                                            
                                                    if frame_name or frame_id:
                                                        driver.switch_to.default_content()
                                                        if frame_name:
                                                            driver.switch_to.frame(frame_name)
                                                        else:
                                                            driver.switch_to.frame(frame_id)
                                                
                                                        try:
                                                            pod_menu = WebDriverWait(driver, 3).until(
                                                                EC.presence_of_element_located((By.XPATH, "//a[contains(text(),'P.O.D.') or contains(text(),'POD')]"))
                                                            )
                                                            if pod_menu.is_displayed():
                                                                onclick_attr = pod_menu.get_attribute('onclick') or ''
                                                                if 'cambiar_contenido' in onclick_attr:
                                                                    print(f"[OK] Menú 'P.O.D.' encontrado en frame '{frame_name or frame_id}'")
                                                                    driver.execute_script("arguments[0].click();", pod_menu)
                                                                    print("[OK] Clic realizado mediante JavaScript")
                                                                    time.sleep(1)
                                                                    break
                                                                else:
                                                                    pod_menu = None
                                                        except:
                                                            continue
                                        
                                                    if pod_menu:
                                                        break
                                                except:
                                                    continue
                                    
                                            if pod_menu:
                                                break
                                        except:
                                            continue
                        
                                if not pod_menu:
                                    print("[WARNING] No se encontró el menú 'P.O.D.'.")
                                    print("   El navegador permanecerá abierto para que puedas revisar manualmente.")
                                else:
                                    # PASO 7: Buscar botón "V" (Ver imagen) y luego descargar todos los PDFs disponibles
                                    print("\n[PASO 7] Buscando botón 'V' (Ver imagen) y descargando archivos PDF...")
                                    time.sleep(1)  # Optimizado: reducido de 3s a 1s
                            
                                    # Crear estructura de carpetas: DESCARGA/fecha/numero_guia/
                                    fecha_descarga = datetime.now().strftime("%Y%m%d")
                                    carpeta_raiz = DESCARGA_DIR
                                    carpeta_fecha = os.path.join(carpeta_raiz, fecha_descarga)
                                    carpeta_guia = os.path.join(carpeta_fecha, guia_prueba)
                            
                                    # Crear carpetas si no existen
                                    os.makedirs(carpeta_guia, exist_ok=True)
                                    print(f"[OK] Carpeta de descarga creada: {carpeta_guia}")
                            
                                    # ============================================================
                                    # PASO 7.1 - Buscar y descargar IMÁGENES con código de guía
                                    # ============================================================
                                    print(f"\n   [PASO 7.1] Buscando imágenes que contengan '{guia_prueba}' en su nombre...")
                                    ventana_principal = driver.current_window_handle
                            
                                    # Buscar imágenes con JS recursivo (rápido, en una sola llamada)
                                    seccion_imagenes = driver.execute_script("""
                                        var imagenes = [];
                                        var hrefsVistos = {};
                                        var numGuia = arguments[0];
                                
                                        function buscar(doc, nivel) {
                                            if (nivel > 10) return;
                                            try {
                                                // Buscar sección "Imágenes disponibles"
                                                var tds = doc.querySelectorAll('td');
                                                var tdSec = null;
                                                for (var i = 0; i < tds.length; i++) {
                                                    var t = (tds[i].textContent || '').trim();
                                                    if (t.indexOf('mágenes disponibles') !== -1 || t.indexOf('Imagenes disponibles') !== -1) {
                                                        tdSec = tds[i]; break;
                                                    }
                                                }
                                                if (tdSec) {
                                                    var tabla = tdSec.closest('table') || tdSec.parentElement;
                                                    var enlaces = tabla.querySelectorAll('a[href]');
                                                    for (var e = 0; e < enlaces.length; e++) {
                                                        var raw = enlaces[e].getAttribute('href') || '';
                                                        if (!raw || hrefsVistos[raw]) continue;
                                                
                                                        // FILTRO: solo si contiene el código de guía en rawHref, texto o parámetros
                                                        var texto = (enlaces[e].textContent || '').trim();
                                                        var contieneGuia = raw.indexOf(numGuia) !== -1 || texto.indexOf(numGuia) !== -1;
                                                
                                                        // Extraer parámetros JS para verificar
                                                        var params = [];
                                                        var jsCode = '';
                                                        if (raw.indexOf('javascript:') === 0) {
                                                            jsCode = raw.substring(11);
                                                            var rx = /['"]([^'"]+)['"]/g, m;
                                                            while ((m = rx.exec(jsCode)) !== null) {
                                                                params.push(m[1]);
                                                                if (m[1].indexOf(numGuia) !== -1) contieneGuia = true;
                                                            }
                                                        }
                                                
                                                        if (!contieneGuia) continue;
                                                        hrefsVistos[raw] = true;
                                                
                                                        imagenes.push({
                                                            rawHref: raw, href: enlaces[e].href || raw,
                                                            texto: texto || 'img_' + imagenes.length,
                                                            jsCode: jsCode, params: params
                                                        });
                                                    }
                                                }
                                                var frames = doc.querySelectorAll('frame, iframe');
                                                for (var f = 0; f < frames.length; f++) {
                                                    try { buscar(frames[f].contentDocument || frames[f].contentWindow.document, nivel+1); } catch(ex) {}
                                                }
                                            } catch(ex) {}
                                        }
                                        buscar(document, 0);
                                        return imagenes;
                                    """, guia_prueba)
                            
                                    if seccion_imagenes and len(seccion_imagenes) > 0:
                                        print(f"      [OK] {len(seccion_imagenes)} imagen(es) con guía '{guia_prueba}' encontrada(s)")
                                
                                        # Preparar cookies y URLs base (una sola vez) - Optimizado con sesión
                                        url_base = '/'.join(driver.current_url.split('/')[:3])
                                        cookies = driver.get_cookies()
                                        cookie_dict = {c['name']: c['value'] for c in cookies}
                                        cookie_str = '; '.join([f"{c['name']}={c['value']}" for c in cookies])
                                        imagenes_descargadas = 0
                                
                                        # Crear sesión requests para mejor rendimiento (reutiliza conexiones)
                                        if requests_available:
                                            session = requests.Session()
                                            session.cookies.update(cookie_dict)
                                            session.headers.update({
                                                'User-Agent': driver.execute_script("return navigator.userAgent;"),
                                                'Referer': driver.current_url
                                            })
                                
                                        def detectar_ext(data, hint=''):
                                            if data[:4] == b'\x89PNG': return '.png'
                                            if data[:2] == b'\xff\xd8': return '.jpg'
                                            if data[:5] == b'%PDF-': return '.pdf'
                                            if data[:4] == b'GIF8': return '.gif'
                                            for e in ['.png','.gif','.jpeg','.pdf','.jpg']:
                                                if e in hint.lower(): return e
                                            return '.jpg'
                                
                                        def descargar_auth(url, ruta):
                                            # Optimizado: usar requests si está disponible (más rápido)
                                            if requests_available:
                                                try:
                                                    response = session.get(url, timeout=15, stream=True)
                                                    if response.status_code == 200:
                                                        data = response.content
                                                        if len(data) > 200:
                                                            with open(ruta, 'wb') as f:
                                                                f.write(data)
                                                            return data
                                                except Exception as e:
                                                    print(f"      [WARNING] Error con requests: {e}, usando urllib como fallback")
                                            # Fallback a urllib
                                            op = urllib.request.build_opener()
                                            op.addheaders = [('Cookie', cookie_str)]
                                            r = op.open(url, timeout=15)
                                            data = r.read()
                                            if len(data) > 200:
                                                with open(ruta, 'wb') as f: f.write(data)
                                                return data
                                            return None
                                
                                        for idx, img_info in enumerate(seccion_imagenes):
                                            try:
                                                rawHref = img_info.get('rawHref', '')
                                                resolvedUrl = img_info.get('href', '')
                                                jsCode = img_info.get('jsCode', '')
                                                params = img_info.get('params', [])
                                                texto_img = img_info.get('texto', f'img_{idx+1}')
                                        
                                                nombre_base = f"img_{idx+1}_{guia_prueba}"
                                                nombre_base = ''.join(c for c in nombre_base if c.isalnum() or c in ('_', '-', '.'))[:100]
                                                imagen_ok = False
                                        
                                                # 1) URL directa
                                                if not imagen_ok and resolvedUrl and not resolvedUrl.startswith('javascript:'):
                                                    try:
                                                        if not resolvedUrl.startswith('http'):
                                                            resolvedUrl = urllib.parse.urljoin(url_base, resolvedUrl)
                                                        tmp = os.path.join(carpeta_guia, nombre_base + '.tmp')
                                                        data = descargar_auth(resolvedUrl, tmp)
                                                        if data and not data[:6].startswith(b'<html') and not data[:15].startswith(b'<!DOCTYPE'):
                                                            ext = detectar_ext(data, resolvedUrl)
                                                            final = os.path.join(carpeta_guia, nombre_base + ext)
                                                            os.rename(tmp, final)
                                                            imagenes_descargadas += 1; imagen_ok = True
                                                        else:
                                                            try: os.remove(tmp)
                                                            except: pass
                                                    except:
                                                        try: os.remove(os.path.join(carpeta_guia, nombre_base + '.tmp'))
                                                        except: pass
                                        
                                                # 2) Interceptar window.open del JS
                                                if not imagen_ok and jsCode:
                                                    try:
                                                        url_int = driver.execute_script("""
                                                            var cap = null;
                                                            function run(doc, n, code) {
                                                                if (n > 10) return null;
                                                                try {
                                                                    var w = doc.defaultView || window;
                                                                    var orig = w.open;
                                                                    w.open = function(u) { cap = u; w.open = orig; return null; };
                                                                    try { (new Function(code)).call(w); } catch(e) {}
                                                                    w.open = orig;
                                                                    if (cap) return cap;
                                                                    var fr = doc.querySelectorAll('frame, iframe');
                                                                    for (var i = 0; i < fr.length; i++) {
                                                                        try { var r = run(fr[i].contentDocument || fr[i].contentWindow.document, n+1, code); if (r) return r; } catch(e) {}
                                                                    }
                                                                } catch(e) {}
                                                                return null;
                                                            }
                                                            return run(document, 0, arguments[0]);
                                                        """, jsCode)
                                                        if url_int:
                                                            if not url_int.startswith('http'):
                                                                url_int = urllib.parse.urljoin(url_base, url_int)
                                                            tmp = os.path.join(carpeta_guia, nombre_base + '.tmp')
                                                            data = descargar_auth(url_int, tmp)
                                                            if data and not data[:15].startswith(b'<!DOCTYPE'):
                                                                ext = detectar_ext(data, url_int)
                                                                final = os.path.join(carpeta_guia, nombre_base + ext)
                                                                os.rename(tmp, final)
                                                                imagenes_descargadas += 1; imagen_ok = True
                                                            else:
                                                                try: os.remove(tmp)
                                                                except: pass
                                                    except:
                                                        pass
                                        
                                                # 3) Click recursivo + nueva ventana
                                                if not imagen_ok and rawHref:
                                                    try:
                                                        vent_act = driver.current_window_handle
                                                        vent_antes = set(driver.window_handles)
                                                        clicked = driver.execute_script("""
                                                            function click(doc, n, h) {
                                                                if (n > 10) return false;
                                                                try {
                                                                    var as = doc.querySelectorAll('a[href]');
                                                                    for (var i = 0; i < as.length; i++) { if (as[i].getAttribute('href') === h) { as[i].click(); return true; } }
                                                                    var fr = doc.querySelectorAll('frame, iframe');
                                                                    for (var f = 0; f < fr.length; f++) { try { if (click(fr[f].contentDocument || fr[f].contentWindow.document, n+1, h)) return true; } catch(e) {} }
                                                                } catch(e) {}
                                                                return false;
                                                            }
                                                            return click(document, 0, arguments[0]);
                                                        """, rawHref)
                                                        if clicked:
                                                            time.sleep(2)
                                                            nuevas = set(driver.window_handles) - vent_antes
                                                            if nuevas:
                                                                driver.switch_to.window(list(nuevas)[0])
                                                                time.sleep(1)
                                                                url_nv = driver.current_url
                                                                # Buscar img o descargar URL directa
                                                                try:
                                                                    imgs_nv = driver.find_elements(By.TAG_NAME, "img")
                                                                    for im in imgs_nv:
                                                                        src = im.get_attribute('src') or ''
                                                                        if src and 'about:' not in src and not src.startswith('data:'):
                                                                            if not src.startswith('http'): src = urllib.parse.urljoin(url_base, src)
                                                                            tmp = os.path.join(carpeta_guia, nombre_base + '.tmp')
                                                                            data = descargar_auth(src, tmp)
                                                                            if data:
                                                                                ext = detectar_ext(data, src)
                                                                                os.rename(tmp, os.path.join(carpeta_guia, nombre_base + ext))
                                                                                imagenes_descargadas += 1; imagen_ok = True; break
                                                                except: pass
                                                                if not imagen_ok and url_nv and not url_nv.startswith('about:'):
                                                                    try:
                                                                        tmp = os.path.join(carpeta_guia, nombre_base + '.tmp')
                                                                        data = descargar_auth(url_nv, tmp)
                                                                        if data and not data[:15].startswith(b'<!DOCTYPE'):
                                                                            ext = detectar_ext(data, url_nv)
                                                                            os.rename(tmp, os.path.join(carpeta_guia, nombre_base + ext))
                                                                            imagenes_descargadas += 1; imagen_ok = True
                                                                        else:
                                                                            try: os.remove(tmp)
                                                                            except: pass
                                                                    except:
                                                                        try: os.remove(os.path.join(carpeta_guia, nombre_base + '.tmp'))
                                                                        except: pass
                                                                try: driver.close()
                                                                except: pass
                                                                driver.switch_to.window(vent_act)
                                                    except:
                                                        try: driver.switch_to.window(ventana_principal)
                                                        except: pass
                                        
                                                status = "[OK]" if imagen_ok else "[WARNING]"
                                                print(f"      {status} Imagen #{idx+1}/{len(seccion_imagenes)}: {texto_img[:50]}")
                                            except Exception as e:
                                                print(f"      [WARNING] Error imagen #{idx+1}: {e}")
                                                try: driver.switch_to.window(ventana_principal)
                                                except: pass
                                
                                        print(f"      [OK] {imagenes_descargadas}/{len(seccion_imagenes)} imagen(es) descargada(s).")
                                    else:
                                        print(f"      [WARNING] No se encontraron imágenes con código de guía '{guia_prueba}'.")
                            
                                    # Asegurar que estamos en la ventana principal antes de buscar PDFs
                                    try:
                                        driver.switch_to.window(ventana_principal)
                                        driver.switch_to.default_content()
                                    except:
                                        pass
                            
                                    # ============================================================
                                    # PASO 7.2 - Buscar y descargar PDFs con código de guía
                                    # ============================================================
                                    print(f"\n   [PASO 7.2] Buscando PDFs que contengan '{guia_prueba}'...")
                                    ventana_principal_pdf = driver.current_window_handle
                            
                                    archivos_antes_pdf = set(os.listdir(carpeta_guia)) if os.path.exists(carpeta_guia) else set()
                                    carpeta_descargas_chrome = DESCARGA_DIR
                                    archivos_descargas_antes = set(os.listdir(carpeta_descargas_chrome)) if os.path.exists(carpeta_descargas_chrome) else set()
                            
                                    try:
                                        driver.switch_to.window(ventana_principal_pdf)
                                        driver.switch_to.default_content()
                                    except: pass
                            
                                    # Búsqueda con JS recursivo (rápido, una sola llamada)
                                    pdfs_js = driver.execute_script("""
                                        var pdfs = [];
                                        var numGuia = arguments[0];
                                        var vistos = {};
                                
                                        function buscar(doc, nivel) {
                                            if (nivel > 10) return;
                                            try {
                                                var tds = doc.querySelectorAll('td');
                                                for (var t = 0; t < tds.length; t++) {
                                                    var txt = (tds[t].textContent || '').trim();
                                                    if (txt.indexOf(numGuia) === -1 || (txt.indexOf('.pdf') === -1 && txt.indexOf('.PDF') === -1)) continue;
                                                    if (vistos[txt]) continue;
                                                    vistos[txt] = true;
                                            
                                                    var fila = tds[t].closest('tr');
                                                    var href = '', rawHref = '', onclick = '', jsCode = '', params = [];
                                            
                                                    // Buscar enlace en td, luego en fila, luego clickeables
                                                    var el = tds[t].querySelector('a') || (fila ? fila.querySelector('a') : null);
                                                    if (el) {
                                                        href = el.href || ''; rawHref = el.getAttribute('href') || '';
                                                        onclick = el.getAttribute('onclick') || '';
                                                    }
                                                    if (!href && !onclick && fila) {
                                                        var cls = fila.querySelectorAll('[onclick]');
                                                        for (var c = 0; c < cls.length; c++) {
                                                            onclick = cls[c].getAttribute('onclick') || '';
                                                            if (onclick) break;
                                                        }
                                                    }
                                            
                                                    // Extraer jsCode y params
                                                    if (rawHref && rawHref.indexOf('javascript:') === 0) jsCode = rawHref.substring(11);
                                                    else if (onclick) jsCode = onclick;
                                                    if (jsCode) {
                                                        var rx = /['"]([^'"]+)['"]/g, m;
                                                        while ((m = rx.exec(jsCode)) !== null) params.push(m[1]);
                                                    }
                                            
                                                    pdfs.push({href:href, rawHref:rawHref, onclick:onclick, jsCode:jsCode, params:params, texto:txt});
                                                }
                                                var frames = doc.querySelectorAll('frame, iframe');
                                                for (var f = 0; f < frames.length; f++) {
                                                    try { buscar(frames[f].contentDocument || frames[f].contentWindow.document, nivel+1); } catch(e) {}
                                                }
                                            } catch(e) {}
                                        }
                                        buscar(document, 0);
                                        return pdfs;
                                    """, guia_prueba)
                            
                                    if pdfs_js and len(pdfs_js) > 0:
                                        print(f"      [OK] {len(pdfs_js)} PDF(s) encontrado(s)")
                                
                                        import re
                                        url_base_pdf = '/'.join(driver.current_url.split('/')[:3])
                                        cookies_pdf = driver.get_cookies()
                                        cookie_dict_pdf = {c['name']: c['value'] for c in cookies_pdf}
                                        cookie_str_pdf = '; '.join([f"{c['name']}={c['value']}" for c in cookies_pdf])
                                
                                        # Crear sesión requests para PDFs (reutiliza conexiones)
                                        if requests_available:
                                            session_pdf = requests.Session()
                                            session_pdf.cookies.update(cookie_dict_pdf)
                                            session_pdf.headers.update({
                                                'User-Agent': driver.execute_script("return navigator.userAgent;"),
                                                'Referer': driver.current_url
                                            })
                                
                                        def dl_pdf(url, ruta):
                                            # Optimizado: usar requests si está disponible (más rápido)
                                            if requests_available:
                                                try:
                                                    response = session_pdf.get(url, timeout=15, stream=True)
                                                    if response.status_code == 200:
                                                        data = response.content
                                                        if len(data) > 100 and data[:5] == b'%PDF-':
                                                            with open(ruta, 'wb') as f:
                                                                f.write(data)
                                                            return True
                                                except Exception as e:
                                                    print(f"      [WARNING] Error con requests: {e}, usando urllib como fallback")
                                            # Fallback a urllib
                                            op = urllib.request.build_opener()
                                            op.addheaders = [('Cookie', cookie_str_pdf)]
                                            data = op.open(url, timeout=15).read()
                                            if len(data) > 100 and data[:5] == b'%PDF-':
                                                with open(ruta, 'wb') as f: f.write(data)
                                                return True
                                            return False
                                
                                        pdfs_ok = 0
                                        for idx, pj in enumerate(pdfs_js):
                                            txt = pj.get('texto', f'PDF_{idx+1}')
                                            href = pj.get('href', '')
                                            jsCode = pj.get('jsCode', '')
                                            onclick = pj.get('onclick', '')
                                            params = pj.get('params', [])
                                    
                                            nom = txt.strip()
                                            if not nom.lower().endswith('.pdf'): nom += '.pdf'
                                            nom = ''.join(c for c in nom if c.isalnum() or c in ('_', '-', '.'))[:150]
                                            ruta = os.path.join(carpeta_guia, nom)
                                            ok = False
                                            vent_antes = set(driver.window_handles)
                                    
                                            # A) URL HTTP directa
                                            if not ok and href and href.startswith('http'):
                                                try: ok = dl_pdf(href, ruta)
                                                except: pass
                                    
                                            # B) Extraer URL del JS
                                            if not ok and (onclick or jsCode):
                                                js_src = onclick or jsCode
                                                urls = re.findall(r"['\"]([^'\"]*\.pdf[^'\"]*)['\"]", js_src, re.IGNORECASE)
                                                if not urls: urls = re.findall(r"['\"]([^'\"]+)['\"]", js_src)
                                                for u in urls:
                                                    try:
                                                        if not u.startswith('http'): u = urllib.parse.urljoin(url_base_pdf, u)
                                                        if dl_pdf(u, ruta): ok = True; break
                                                    except: continue
                                        
                                                # Ejecutar JS recursivamente + nueva pestaña
                                                if not ok:
                                                    try:
                                                        driver.execute_script("""
                                                            function run(doc,n,c){if(n>10)return;try{(new Function(c)).call(doc.defaultView||window)}catch(e){}
                                                            var f=doc.querySelectorAll('frame,iframe');for(var i=0;i<f.length;i++){try{run(f[i].contentDocument||f[i].contentWindow.document,n+1,c)}catch(e){}}}
                                                            run(document,0,arguments[0]);
                                                        """, js_src)
                                                        time.sleep(2)
                                                        nuevas = set(driver.window_handles) - vent_antes
                                                        if nuevas:
                                                            driver.switch_to.window(list(nuevas)[0])
                                                            time.sleep(1)
                                                            try: ok = dl_pdf(driver.current_url, ruta)
                                                            except:
                                                                try:
                                                                    urllib.request.urlretrieve(driver.current_url, ruta)
                                                                    ok = os.path.getsize(ruta) > 100
                                                                except: pass
                                                            try: driver.close()
                                                            except: pass
                                                            driver.switch_to.window(ventana_principal_pdf)
                                                    except: pass
                                    
                                            # C) Clic recursivo en frames (FALLBACK)
                                            if not ok:
                                                try:
                                                    res = driver.execute_script("""
                                                        var tb=arguments[0], ng=arguments[1];
                                                        function click(doc,n){if(n>10)return null;try{
                                                            var tds=doc.querySelectorAll('td');
                                                            for(var t=0;t<tds.length;t++){
                                                                var tx=(tds[t].textContent||'').trim();
                                                                if(tx===tb||(tx.indexOf(ng)!==-1&&tx.indexOf('.pdf')!==-1)){
                                                                    var fila=tds[t].closest('tr');if(!fila)continue;
                                                                    var cls=fila.querySelectorAll('a,button,[onclick]');
                                                                    for(var c=0;c<cls.length;c++){if(cls[c].offsetParent!==null){cls[c].click();return{ok:true}}}
                                                                    tds[t].click();return{ok:true};
                                                                }
                                                            }
                                                            var fr=doc.querySelectorAll('frame,iframe');
                                                            for(var f=0;f<fr.length;f++){try{var r=click(fr[f].contentDocument||fr[f].contentWindow.document,n+1);if(r)return r}catch(e){}}
                                                        }catch(e){}return null}
                                                        return click(document,0);
                                                    """, txt, guia_prueba)
                                                    if res and res.get('ok'):
                                                        time.sleep(3)
                                                        nuevas = set(driver.window_handles) - vent_antes
                                                        if nuevas:
                                                            driver.switch_to.window(list(nuevas)[0])
                                                            time.sleep(1)
                                                            try: ok = dl_pdf(driver.current_url, ruta)
                                                            except:
                                                                try:
                                                                    urllib.request.urlretrieve(driver.current_url, ruta)
                                                                    ok = os.path.getsize(ruta) > 100
                                                                except: pass
                                                            try: driver.close()
                                                            except: pass
                                                            driver.switch_to.window(ventana_principal_pdf)
                                                        else:
                                                            # Verificar descarga automática
                                                            time.sleep(3)
                                                            import shutil
                                                            desc_new = set(os.listdir(carpeta_descargas_chrome)) - archivos_descargas_antes
                                                            for d in desc_new:
                                                                if d.lower().endswith('.pdf') and guia_prueba in d:
                                                                    try:
                                                                        shutil.move(os.path.join(carpeta_descargas_chrome, d), os.path.join(carpeta_guia, d))
                                                                        ok = True; archivos_descargas_antes.add(d)
                                                                    except: ok = True
                                                except: pass
                                    
                                            status = "[OK]" if ok else "[WARNING]"
                                            if ok: pdfs_ok += 1
                                            print(f"      {status} PDF #{idx+1}/{len(pdfs_js)}: {txt[:60]}")
                                
                                        # Resumen
                                        archivos_finales = set(os.listdir(carpeta_guia)) if os.path.exists(carpeta_guia) else set()
                                        nuevos_pdfs = [f for f in (archivos_finales - archivos_antes_pdf) if f.lower().endswith('.pdf')]
                                        if nuevos_pdfs:
                                            print(f"      [OK] {len(nuevos_pdfs)} PDF(s) descargado(s) en: {carpeta_guia}")
                                            for p in nuevos_pdfs: print(f"         - {p}")
                                        else:
                                            print(f"      [WARNING] No se pudieron descargar PDFs.")
                                    else:
                                        print(f"      [WARNING] No se encontraron PDFs con '{guia_prueba}'.")
                            
                                    # Volver a la ventana principal al finalizar
                                    try:
                                        driver.switch_to.window(ventana_principal_pdf)
                                        driver.switch_to.default_content()
                                    except:
                                        pass

                            guia_exitosa = True
                        except Exception as e_guia:
                            print(f"\n[FAIL] Error procesando guía {guia_prueba}: {e_guia}")
                            import traceback
                            traceback.print_exc()
                        finally:
                            # === ACTUALIZAR EXCEL (siempre se ejecuta, incluso con continue) ===
                            try:
                                if guia_exitosa:
                                    # Verificar si se descargaron archivos PDF
                                    fecha_desc = datetime.now().strftime("%Y%m%d")
                                    carpeta_check = os.path.join(DESCARGA_DIR, fecha_desc, guia_prueba)
                                    if os.path.exists(carpeta_check):
                                        archivos_en_carpeta = os.listdir(carpeta_check)
                                        pdfs_check = [f for f in archivos_en_carpeta if f.lower().endswith(".pdf")]
                                        imgs_check = [f for f in archivos_en_carpeta if f.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tif", ".tiff"))]
                                        if pdfs_check or imgs_check:
                                            ws_guias.cell(row=fila_excel, column=2).value = "OK"
                                            guias_ok += 1
                                            print(f"\n[OK] GUÍA {guia_prueba} → OK ({len(pdfs_check)} PDFs, {len(imgs_check)} imágenes)")
                                        else:
                                            ws_guias.cell(row=fila_excel, column=2).value = "SIN ARCHIVOS"
                                            guias_error += 1
                                            print(f"\n[WARNING] GUÍA {guia_prueba} → SIN ARCHIVOS descargados")
                                    else:
                                        ws_guias.cell(row=fila_excel, column=2).value = "SIN DATOS"
                                        guias_error += 1
                                        print(f"\n[WARNING] GUÍA {guia_prueba} → SIN DATOS (carpeta no creada)")
                                else:
                                    ws_guias.cell(row=fila_excel, column=2).value = "ERROR"
                                    guias_error += 1
                                    print(f"\n[FAIL] GUÍA {guia_prueba} → ERROR")
                                try:
                                    wb_guias.save(ruta_excel)
                                except PermissionError:
                                    wb_guias.save(ruta_excel_backup)
                                    print(f"   (Guardado en copia: {ruta_excel_backup})")
                            except Exception as e_excel:
                                print(f"[WARNING] Error actualizando Excel: {e_excel}")

                            # === NAVEGAR DE VUELTA AL FORMULARIO PARA LA SIGUIENTE GUÍA ===
                            if idx_guia < total_guias - 1:
                                print("\nNavegando de vuelta al formulario de búsqueda...")
                                try:
                                    # Cerrar pestañas/ventanas extras (quedarse solo con la principal)
                                    ventanas_actuales = driver.window_handles
                                    if len(ventanas_actuales) > 1:
                                        ventana_principal_loop = ventanas_actuales[0]
                                        for v_extra in ventanas_actuales[1:]:
                                            try:
                                                driver.switch_to.window(v_extra)
                                                driver.close()
                                            except:
                                                pass
                                        driver.switch_to.window(ventana_principal_loop)
                                    driver.switch_to.default_content()
                                    time.sleep(1)

                                    # Verificar si la página sigue siendo el frameset
                                    # Si hay error POST resubmission, el frameset se destruye
                                    pagina_ok = False
                                    try:
                                        frames_check = driver.find_elements(By.TAG_NAME, "frame")
                                        if len(frames_check) > 0:
                                            pagina_ok = True
                                            print("   [OK] Frameset detectado, página OK")
                                    except:
                                        pass
                                    
                                    if not pagina_ok:
                                        # El frameset se destruyó (error POST). Navegar a la URL principal.
                                        print("   [WARNING] Frameset no detectado. Recuperando página principal...")
                                        try:
                                            driver.get(url_principal_app)
                                            time.sleep(5)
                                            driver.switch_to.default_content()
                                            # Verificar de nuevo
                                            frames_check = driver.find_elements(By.TAG_NAME, "frame")
                                            if len(frames_check) > 0:
                                                pagina_ok = True
                                                print("   [OK] Página principal recuperada exitosamente")
                                            else:
                                                print("   [WARNING] No se pudo recuperar el frameset")
                                        except Exception as e_rec:
                                            print(f"   [WARNING] Error recuperando página: {e_rec}")

                                    if pagina_ok:
                                        # RE-NAVEGAR al 7.8 usando el buscador del menú
                                        print("   Re-navegando al menú 7.8 Consulta a Histórico...")
                                        
                                        # 1. Ir al frame "menu" y buscar funcionalidad_codigo
                                        campo_menu = None
                                        try:
                                            driver.switch_to.frame("menu")
                                            campo_menu = WebDriverWait(driver, 5).until(
                                                EC.presence_of_element_located((By.NAME, "funcionalidad_codigo"))
                                            )
                                        except:
                                            driver.switch_to.default_content()
                                            # Buscar el frame menu por otros medios
                                            try:
                                                frames = driver.find_elements(By.TAG_NAME, "frame")
                                                for fr in frames:
                                                    try:
                                                        fr_name = fr.get_attribute("name") or ""
                                                        if fr_name == "menu" or "menu" in fr_name.lower():
                                                            driver.switch_to.frame(fr)
                                                            campo_menu = WebDriverWait(driver, 5).until(
                                                                EC.presence_of_element_located((By.NAME, "funcionalidad_codigo"))
                                                            )
                                                            break
                                                    except:
                                                        driver.switch_to.default_content()
                                                        continue
                                            except:
                                                pass
                                        
                                        if campo_menu:
                                            # 2. Limpiar, escribir "7.8" y presionar Enter
                                            try:
                                                campo_menu.clear()
                                                time.sleep(0.3)
                                                campo_menu.send_keys("7.8")
                                                time.sleep(1)  # Esperar a que aparezca autocomplete
                                                campo_menu.send_keys(Keys.ENTER)
                                                print("   [OK] '7.8' escrito y Enter enviado en el menú")
                                                time.sleep(4)  # Esperar a que cargue el formulario
                                                
                                                # 3. Verificar que el formulario se cargó (buscar campo nenvio)
                                                driver.switch_to.default_content()
                                                formulario_cargado = driver.execute_script("""
                                                    function buscarNenvio(doc, nivel) {
                                                        if (nivel > 10) return false;
                                                        try {
                                                            var campo = doc.querySelector('input[name="nenvio"]');
                                                            if (campo && campo.offsetParent !== null) return true;
                                                            var frames = doc.querySelectorAll('frame, iframe');
                                                            for (var i = 0; i < frames.length; i++) {
                                                                try {
                                                                    if (buscarNenvio(frames[i].contentDocument, nivel + 1)) return true;
                                                                } catch(e) {}
                                                            }
                                                        } catch(e) {}
                                                        return false;
                                                    }
                                                    return buscarNenvio(document, 0);
                                                """)
                                                
                                                if formulario_cargado:
                                                    print("   [OK] Formulario 7.8 cargado exitosamente (campo nenvio detectado)")
                                                else:
                                                    # Tal vez Enter no seleccionó el resultado. Buscar y clickear manualmente.
                                                    print("   [WARNING] Formulario no detectado. Buscando resultado '7.8' para clickear...")
                                                    driver.switch_to.default_content()
                                                    resultado_click = driver.execute_script("""
                                                        function buscarYClick78(doc, nivel) {
                                                            if (nivel > 10) return false;
                                                            try {
                                                                var elementos = doc.querySelectorAll('a, span, li, div, td');
                                                                for (var i = 0; i < elementos.length; i++) {
                                                                    var texto = (elementos[i].textContent || '').trim();
                                                                    if (texto.indexOf('7.8') !== -1 && elementos[i].offsetParent !== null) {
                                                                        if (texto.indexOf('Hist') !== -1 || texto.indexOf('Consulta') !== -1 || texto === '7.8') {
                                                                            elementos[i].click();
                                                                            return true;
                                                                        }
                                                                    }
                                                                }
                                                                var frames = doc.querySelectorAll('frame, iframe');
                                                                for (var f = 0; f < frames.length; f++) {
                                                                    try {
                                                                        if (buscarYClick78(frames[f].contentDocument, nivel + 1)) return true;
                                                                    } catch(e) {}
                                                                }
                                                            } catch(e) {}
                                                            return false;
                                                        }
                                                        return buscarYClick78(document, 0);
                                                    """)
                                                    if resultado_click:
                                                        print("   [OK] Clic en resultado '7.8' realizado")
                                                        time.sleep(4)
                                                    else:
                                                        print("   [WARNING] No se encontró resultado '7.8'. Continuando de todas formas...")
                                                    driver.switch_to.default_content()
                                            except Exception as e_menu:
                                                print(f"   [WARNING] Error al usar el buscador del menú: {e_menu}")
                                                driver.switch_to.default_content()
                                        else:
                                            print("   [WARNING] No se encontró el campo del menú")
                                            driver.switch_to.default_content()
                                    else:
                                        print("   [WARNING] No se pudo recuperar la página. Intentando continuar...")
                                        driver.switch_to.default_content()
                                except Exception as e_nav:
                                    print(f"[WARNING] Error al navegar de vuelta: {e_nav}")
                                    try:
                                        driver.switch_to.default_content()
                                    except:
                                        pass

                    # === RESUMEN FINAL DE DESCARGA MASIVA ===
                    print(f"\n{"=" * 60}")
                    print("RESUMEN DESCARGA MASIVA DE GUÍAS")
                    print(f"{"=" * 60}")
                    print(f"Total guías procesadas: {total_guias}")
                    print(f"Guías exitosas (OK):    {guias_ok}")
                    print(f"Guías con error:        {guias_error}")
                    print(f"\nArchivo Excel actualizado: {ruta_excel}")

                except Exception as e:
                    print(f"\n[FAIL] Error durante la navegación al menú 7.8 o consulta de POD: {e}")
                    import traceback
                    print("\nDetalles del error:")
                    traceback.print_exc()
            else:
                # Aunque no se detectó claramente el login, intentar continuar si no hay errores
                print("\n[WARNING] No se detectó cambio claro en la URL o elementos de la página principal")
                print("Intentando continuar de todas formas (puede que el login haya sido exitoso)...")
                
                # Intentar navegar al menú de todas formas
                print("\n" + "=" * 60)
                print("NAVEGACIÓN A MENÚ 7. SAC -> 7.8 Consulta a historico")
                print("=" * 60)
                # guia_prueba se asigna desde el Excel en el loop de descarga masiva

                try:
                    # Esperar un poco más
                    print("Esperando a que la página cargue completamente...")
                    time.sleep(5)
                    
                    # Intentar buscar el menú 7.SAC de todas formas
                    print("\n[PASO 1] Buscando menú principal '7.SAC'...")
                    menu_sac = None
                    
                    try:
                        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "pui-button-text")))
                        print("[OK] Elementos con clase 'pui-button-text' encontrados")
                    except TimeoutException:
                        print("[WARNING] Esperando elementos del menú...")
                    
                    try:
                        elementos_pui = driver.find_elements(By.CLASS_NAME, "pui-button-text")
                        print(f"   Encontrados {len(elementos_pui)} elementos con clase 'pui-button-text'")
                        
                        for i, elem in enumerate(elementos_pui):
                            try:
                                texto = elem.text.strip()
                                print(f"   Elemento {i+1}: texto = '{texto}'")
                                if "7" in texto and "SAC" in texto.upper():
                                    print(f"   [OK] Coincidencia encontrada: '{texto}'")
                                    if elem.tag_name.lower() == "span":
                                        try:
                                            padre = elem.find_element(By.XPATH, "./parent::*")
                                            if padre:
                                                menu_sac = padre
                                                print(f"[OK] Menú '7.SAC' encontrado (elemento padre del span)")
                                                break
                                        except Exception:
                                            menu_sac = elem
                                            print(f"[OK] Menú '7.SAC' encontrado (span directamente)")
                                            break
                                    else:
                                        menu_sac = elem
                                        print(f"[OK] Menú '7.SAC' encontrado")
                                        break
                            except Exception:
                                continue
                    except Exception as e:
                        print(f"Error al buscar elementos: {e}")
                    
                    if menu_sac:
                        try:
                            driver.execute_script("arguments[0].scrollIntoView(true);", menu_sac)
                            time.sleep(0.5)
                            menu_sac.click()
                            print("[OK] Clic realizado en el menú '7.SAC'")
                            time.sleep(2)
                        except Exception as e:
                            try:
                                driver.execute_script("arguments[0].click();", menu_sac)
                                print("[OK] Clic realizado mediante JavaScript en el menú '7.SAC'")
                                time.sleep(2)
                            except Exception as e2:
                                print(f"Error al hacer clic: {e2}")
                    else:
                        print("[WARNING] No se pudo encontrar el menú '7.SAC'.")
                        print("   Revisa manualmente en el navegador si el login fue exitoso.")
                        
                except Exception as e:
                    print(f"\n[FAIL] Error al intentar navegar: {e}")
                    print("Revisa manualmente el estado del login en el navegador.")

        print("\n" + "=" * 60)
        print("El navegador permanecerá abierto para que puedas verificar.")
        print("=" * 60)
        safe_input("\nPulsa ENTER para cerrar el navegador...")
        
    except KeyboardInterrupt:
        print("\n\nOperación cancelada por el usuario.")
        
    except Exception as e:
        print(f"\n[FAIL] Error inesperado: {e}")
        print(f"Tipo: {type(e).__name__}")
        import traceback
        print("\nDetalles:")
        traceback.print_exc()
        safe_input("\nPulsa ENTER para salir...")
        
    finally:
        if driver:
            try:
                driver.quit()
                print("\n[OK] Navegador cerrado correctamente.")
            except Exception:
                pass

if __name__ == "__main__":
    main()
